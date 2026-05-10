"""결정적 분류 엔진 — 동일 입력 10회 반복 → 100% 동일 출력."""
from __future__ import annotations

import re
from pathlib import Path
from typing import Optional, Union

try:
    import pandas as pd
    _PANDAS_OK = True
except ImportError:
    _PANDAS_OK = False

from .account_dict import match_account, ACCOUNT_BY_NAME, SECTION_ORDER
from .ft_classifier import ft_classify

# ── 헤더 자동 감지 키워드 ────────────────────────────────────────────────────
HEADER_KEYWORDS: frozenset[str] = frozenset({
    '거래일', '거래일시', '거래일자', '일자', '일시',
    '거래내용', '적요', '내역', '비고', '메모',
    '출금', '출금액', '출금금액', '지급',
    '입금', '입금액', '입금금액', '금액', '거래금액',
    '잔액', '거래후잔액', '거래후 잔액',
    '거래처', '상대처', '받는분', '보내는분', '상대계좌예금주명',
})

# ── 컬럼 후보 (우선순위순) ─────────────────────────────────────────────────
_DESC_CANDIDATES = [
    "적요", "내용", "거래내용", "거래내역", "description", "memo", "摘要", "비고", "내역",
    "거래적요", "이체메모", "출금내역", "입금내역",
]
_VENDOR_CANDIDATES = [
    "거래처", "상호", "vendor", "payee", "대상", "상대방", "입금처", "출금처",
    "보낸분/받는분", "받는분", "보내는분", "상대계좌예금주명", "상대처",
]
_AMOUNT_CANDIDATES = [
    "금액", "출금액", "입금액", "amount", "출금", "입금", "거래금액", "변동금액",
]
_WITHDRAWAL_CANDIDATES = [
    "출금", "출금액", "출금금액", "지급", "지급액",
]
_DEPOSIT_CANDIDATES = [
    "입금", "입금액", "입금금액",
]
_DATE_CANDIDATES = [
    "거래일", "거래일시", "거래일자", "일자", "일시", "날짜",
]
_TIME_CANDIDATES = [
    "거래시간", "시간",
]
# 2차 적요 컬럼 — 1차 desc_col에 텍스트 결합 (우리은행 메모, 농협 거래기록사항 등)
_MEMO_CANDIDATES = [
    "거래기록사항",  # 농협: 거래내용 세부 기록
    "메모",
    "이체메모",
]

# classify_df가 내부적으로 추가하는 컬럼 — orig_cols에서 제외
_INTERNAL_COLS = {"분류과목", "신뢰도", "_amt", "_datetime"}

# [분류결과] 시트 컬럼 순서: [번호] → [분류과목, 신뢰도] → [아래 순서] → [나머지]
_RESULT_COL_BEFORE_CLASS = ["번호"]
_RESULT_COL_AFTER_CLASS = ["거래일시", "출금", "입금", "거래내용", "상대계좌예금주명"]
_RESULT_DROP_COLS = frozenset([
    "거래후잔액", "상대계좌번호", "상대은행", "메모", "거래구분", "수표어음금액", "CMS코드",
])


def _normalize_col(s: str) -> str:
    """'(원)' 제거 + 공백/유니코드 공백 제거."""
    s = re.sub(r'\(원\)', '', str(s))
    s = re.sub(r'[\s　 ]+', '', s)
    return s


def _build_result_output_cols(orig_cols: "list[str]") -> "tuple[list[str], list[str]]":
    """[분류결과] 출력 컬럼을 (분류과목/신뢰도 앞, 뒤) 두 리스트로 분할."""
    orig_set = set(orig_cols)
    before = [c for c in _RESULT_COL_BEFORE_CLASS if c in orig_set]
    after = [c for c in _RESULT_COL_AFTER_CLASS if c in orig_set]
    all_mentioned = set(_RESULT_COL_BEFORE_CLASS) | set(_RESULT_COL_AFTER_CLASS) | _RESULT_DROP_COLS
    remaining = [c for c in orig_cols if c not in all_mentioned]
    return before, after + remaining


def _detect_col(columns: list[str], candidates: list[str]) -> str | None:
    """컬럼 목록에서 후보 중 첫 번째 일치 컬럼 반환 (정규화 매칭 포함)."""
    normalized_map = {_normalize_col(c): c for c in columns}
    for cand in candidates:
        if cand in columns:
            return cand
        norm_cand = _normalize_col(cand)
        if norm_cand in normalized_map:
            return normalized_map[norm_cand]
    return None


def _detect_header_row(path: Path, max_check: int = 12) -> int:
    """xlsx 첫 max_check 행 스캔 → HEADER_KEYWORDS 매칭 점수 최고 행(0-based) 반환."""
    try:
        import openpyxl
    except ImportError:
        return 0

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    best_row = 0
    best_score = 0.0

    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=max_check, values_only=True)):
        score = 0.0
        for cell in row:
            if cell is None:
                continue
            cell_str = _normalize_col(str(cell).strip())
            # Exact match after normalization
            if cell_str in {_normalize_col(k) for k in HEADER_KEYWORDS}:
                score += 1.0
            else:
                # Substring match (partial keyword in cell)
                for kw in HEADER_KEYWORDS:
                    if _normalize_col(kw) in cell_str:
                        score += 0.5
                        break
        if score > best_score:
            best_score = score
            best_row = i

    wb.close()
    return best_row


def _parse_numeric(s: object) -> float:
    """콤마/￦/원/공백 포함 문자열 → float. 변환 불가 시 0.0."""
    cleaned = re.sub(r"[,￦원\s]", "", str(s))
    try:
        return float(cleaned) if cleaned else 0.0
    except ValueError:
        return 0.0


def _read_file(path: Union[str, Path]) -> "pd.DataFrame":
    """xlsx/csv/xls 자동 감지 후 DataFrame 로드 (xlsx는 헤더 자동 감지)."""
    if not _PANDAS_OK:
        raise ImportError("pandas가 설치되지 않았습니다. pip install pandas openpyxl")
    p = Path(path)
    suffix = p.suffix.lower()
    if suffix == ".xlsx":
        header_row = _detect_header_row(p)
        return pd.read_excel(p, header=header_row, dtype=str).fillna("")
    elif suffix == ".xls":
        return pd.read_excel(p, header=0, dtype=str).fillna("")
    elif suffix == ".csv":
        for enc in ("utf-8-sig", "utf-8", "cp949", "euc-kr"):
            try:
                return pd.read_csv(p, dtype=str, encoding=enc).fillna("")
            except UnicodeDecodeError:
                continue
        raise ValueError(f"CSV 인코딩을 감지할 수 없습니다: {p}")
    else:
        raise ValueError(f"지원하지 않는 파일 형식: {suffix} (지원: .xlsx .xls .csv)")


def _build_classify_text(
    row: "pd.Series",
    desc_col: str | None,
    vendor_col: str | None,
    memo_col: str | None = None,
) -> tuple[str, str]:
    """적요 + 2차메모 결합 + 거래처 추출 → (description, vendor) 반환.

    거래처(vendor) 텍스트도 description에 포함: KB처럼 적요가 '체크카드' 등
    은행 코드이고 실제 상호명이 보낸분/받는분에만 있는 경우 키워드 매칭 활성화.
    """
    parts = []
    if desc_col:
        v = str(row[desc_col]).strip()
        if v and v.lower() != "nan":
            parts.append(v)
    if memo_col:
        v = str(row[memo_col]).strip()
        if v and v.lower() != "nan":
            parts.append(v)
    vendor = ""
    if vendor_col:
        v = str(row[vendor_col]).strip()
        if v and v.lower() != "nan":
            vendor = v
            parts.append(v)
    desc = " ".join(parts)
    return desc, vendor


def classify_df(df: "pd.DataFrame") -> "pd.DataFrame":
    """DataFrame에 [분류과목, 신뢰도] 컬럼을 추가해 반환.

    - 결정적: 동일 입력 → 동일 출력 (랜덤 없음)
    - 미분류 항목은 분류과목="미분류" 신뢰도=0.0
    - 입출금 분리 컬럼 자동 감지 → _amt 통합
    - 농협식 거래일자+거래시간 분리 → _datetime 결합
    - 2차 메모 컬럼 결합 (우리은행 메모, 농협 거래기록사항 등)
    """
    if not _PANDAS_OK:
        raise ImportError("pandas가 설치되지 않았습니다.")

    df = df.copy()
    columns = list(df.columns)

    desc_col = _detect_col(columns, _DESC_CANDIDATES)
    vendor_col = _detect_col(columns, _VENDOR_CANDIDATES)
    memo_col = _detect_col(columns, _MEMO_CANDIDATES)
    # 동일 컬럼 중복 방지
    if memo_col == desc_col:
        memo_col = None

    # 농협: 거래일자 + 거래시간 분리 → _datetime 결합
    date_col = _detect_col(columns, _DATE_CANDIDATES)
    time_col = _detect_col(columns, _TIME_CANDIDATES)
    if date_col and time_col and date_col != time_col:
        df["_datetime"] = df[date_col].astype(str) + " " + df[time_col].astype(str)

    # 입출금 분리 컬럼 → _amt = 입금 - 출금
    withdrawal_col = _detect_col(columns, _WITHDRAWAL_CANDIDATES)
    deposit_col = _detect_col(columns, _DEPOSIT_CANDIDATES)
    if withdrawal_col and deposit_col and withdrawal_col != deposit_col:
        df["_amt"] = df.apply(
            lambda r: _parse_numeric(r[deposit_col]) - _parse_numeric(r[withdrawal_col]),
            axis=1,
        )

    labels: list[str] = []
    confs: list[float] = []

    for _, row in df.iterrows():
        desc, vendor = _build_classify_text(row, desc_col, vendor_col, memo_col)
        amt = float(row["_amt"]) if "_amt" in row.index else 0.0
        if "_amt" in row.index:
            direction: Optional[str] = "입금" if amt > 0 else ("출금" if amt < 0 else None)
        else:
            direction = None
        result = ft_classify(desc, vendor, abs(amt), direction)
        labels.append(result.category)
        confs.append(result.confidence)

    df["분류과목"] = labels
    df["신뢰도"] = confs

    # 계정과목 부호 강제: 비용 계정(_amt > 0)은 음수로 정정
    if "_amt" in df.columns:
        _sign_map = {a.name: a.sign for a in ACCOUNT_BY_NAME.values()}
        _expense_mask = (
            df["분류과목"].map(lambda n: _sign_map.get(n, "+")).eq("-")
            & (df["_amt"] > 0)
        )
        df.loc[_expense_mask, "_amt"] = -df.loc[_expense_mask, "_amt"].abs()

    return df


def classify_file(path: Union[str, Path]) -> "pd.DataFrame":
    """파일 경로 → 분류 결과 DataFrame."""
    df = _read_file(path)
    return classify_df(df)


def save_classified(df: "pd.DataFrame", out_path: Union[str, Path]) -> None:
    """분류 결과를 3시트 xlsx로 저장.

    [분류결과] 원본 컬럼 + 분류과목 + 신뢰도(정수%)  — 분류 성공 행만
    [요약]    분류과목 | 건수 | 합계 | 평균
    [미분류]  원본 컬럼만  — 분류 실패 행
    """
    if not _PANDAS_OK:
        raise ImportError("pandas가 설치되지 않았습니다.")
    try:
        import openpyxl
        from openpyxl.styles import Font
    except ImportError as exc:
        raise RuntimeError("openpyxl 미설치 — pip install openpyxl") from exc

    try:
        from openpyxl.styles import PatternFill, Border, Side
    except ImportError:
        PatternFill = Border = Side = None  # type: ignore[assignment,misc]

    orig_cols = [c for c in df.columns if c not in _INTERNAL_COLS]
    df_ok = df[df["분류과목"] != "미분류"].copy()
    df_ng = df[df["분류과목"] == "미분류"][orig_cols].copy()
    bold = Font(bold=True)
    header_fill = (
        PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
        if PatternFill else None
    )
    thick_top = Border(top=Side(border_style="medium")) if Border and Side else None

    wb = openpyxl.Workbook()

    # ── [요약] — 활성 시트(첫 번째) ─────────────────────────────────────────
    ws2 = wb.active
    ws2.title = "요약"
    ws2.append(["분류과목", "건수", "합계금액", "비율"])
    for cell in ws2[1]:
        cell.font = bold
        if header_fill:
            cell.fill = header_fill

    if not df_ok.empty:
        # _amt: classify_df가 입출금 분리 컬럼에서 미리 계산한 값 우선 사용
        if "_amt" in df_ok.columns:
            grp = df_ok.groupby("분류과목")["_amt"].agg(["count", "sum"]).reset_index()
        else:
            amount_col = _detect_col(list(df.columns), _AMOUNT_CANDIDATES)
            if amount_col:
                cleaned = df_ok[amount_col].astype(str).str.replace(
                    r"[,￦원\s]", "", regex=True
                )
                df_ok["_amt"] = pd.to_numeric(cleaned, errors="coerce").fillna(0)
                grp = df_ok.groupby("분류과목")["_amt"].agg(["count", "sum"]).reset_index()
            else:
                grp = df_ok.groupby("분류과목").size().reset_index(name="count")
                grp["sum"] = 0

        # 재무제표 섹션 순서 + 절댓값 내림차순 정렬
        grp["_section_rank"] = grp["분류과목"].map(
            lambda n: SECTION_ORDER.get(
                getattr(ACCOUNT_BY_NAME.get(n), "section", "other"), 5
            )
        )
        grp = grp.sort_values(
            ["_section_rank", "sum"],
            ascending=[True, False],
            key=lambda s: s.abs() if s.name == "sum" else s,
        ).drop(columns=["_section_rank"])

        t_cnt = int(grp["count"].sum())
        t_sum = int(grp["sum"].sum())
        data_start_row = 2
        for i, (_, r) in enumerate(grp.iterrows()):
            cnt = int(r["count"])
            amt = int(r["sum"])
            ratio = f"{cnt / t_cnt * 100:.1f}%" if t_cnt else "0%"
            ws2.append([r["분류과목"], cnt, amt, ratio])
            ws2.cell(row=data_start_row + i, column=3).number_format = '#,##0"원"'
        totals_row = data_start_row + len(grp)
        ws2.append(["총계", t_cnt, t_sum, "100%"])
        ws2.cell(row=totals_row, column=3).number_format = '#,##0"원"'
        # 합계 행 굵은 상단 테두리 + 볼드
        if thick_top:
            for col_idx in range(1, 5):
                cell = ws2.cell(row=totals_row, column=col_idx)
                cell.border = thick_top
                cell.font = bold

    # ── [분류결과] — 두 번째 시트 ────────────────────────────────────────────
    # 재무제표 섹션 순서 + 절댓값 내림차순 정렬
    df_ok = df_ok.copy()
    df_ok["_section_rank"] = df_ok["분류과목"].map(
        lambda n: SECTION_ORDER.get(
            getattr(ACCOUNT_BY_NAME.get(n), "section", "other"), 5
        )
    )
    if "_amt" in df_ok.columns:
        df_ok["_abs_amt"] = df_ok["_amt"].abs()
        df_ok = df_ok.sort_values(["_section_rank", "_abs_amt"], ascending=[True, False])
        df_ok = df_ok.drop(columns=["_section_rank", "_abs_amt"])
    else:
        df_ok = df_ok.sort_values("_section_rank").drop(columns=["_section_rank"])

    ws1 = wb.create_sheet("분류결과")
    before_cols, after_cols = _build_result_output_cols(orig_cols)
    header = before_cols + ["분류과목", "신뢰도"] + after_cols
    ws1.append(header)
    for cell in ws1[1]:
        cell.font = bold
    for _, row in df_ok.iterrows():
        vals = [row[c] for c in before_cols]
        vals.append(row["분류과목"])
        vals.append(f"{int(round(float(row['신뢰도']) * 100))}%")
        vals.extend(row[c] for c in after_cols)
        ws1.append(vals)

    # ── [미분류] — 세 번째 시트 ──────────────────────────────────────────────
    ws3 = wb.create_sheet("미분류")
    ws3.append(orig_cols)
    for cell in ws3[1]:
        cell.font = bold
    for _, row in df_ng.iterrows():
        ws3.append([row[c] for c in orig_cols])

    wb.save(str(out_path))
