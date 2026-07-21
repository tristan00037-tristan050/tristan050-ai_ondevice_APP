#!/usr/bin/env python3
"""Independent, private-key-free verifier executable for Box5 A4 evidence.

The file is executable under ``python -I`` and contains its own parser,
canonicalizer, graph oracle, fee checks, and DLP rules.  Importing the producer
package is prohibited by construction and checked again in release tests.  It
emits an unsigned verified observation.  Native authority code owns the key,
anti-replay state, receipt assembly, and the only attestation operation.
"""

from __future__ import annotations

import argparse
import base64
import csv
import hashlib
import hmac
import io
import itertools
import json
import os
import re
import stat
import struct
import sys
import tempfile
import unicodedata
import uuid
from dataclasses import dataclass
from datetime import datetime, timezone
from functools import lru_cache
from pathlib import Path
from typing import Any, Mapping, NoReturn, Sequence
from zoneinfo import ZoneInfo

from cryptography.exceptions import InvalidSignature
from cryptography.hazmat.primitives.asymmetric.ed25519 import Ed25519PublicKey

if __package__ in {None, ""}:
    sys.path.insert(0, str(Path(__file__).resolve().parents[2]))

from butler_pc_core.a4_verifier.canonical import sha256_hex
from butler_pc_core.a4_verifier.contracts import (
    MAX_REQUEST_BYTES,
    MAX_RESPONSE_BYTES,
    MAX_SOURCE_BYTES,
    OBSERVATION_SCHEMA,
    PROTOCOL_VERSION,
    VERIFY_REQUEST_SCHEMA,
    AuthorityChallenge,
    request_digest,
)
from butler_pc_core.a4_verifier.errors import public_error_code
from butler_pc_core.a4_verifier.receipt import verified_observation


FILES = frozenset(
    {
        "run_request.json",
        "input_receipt.json",
        "policy_receipt.json",
        "account_snapshot_receipt.json",
        "graph_observation.json",
        "classification_receipt.json",
        "runtime_invariants.json",
        "dlp_receipt.json",
        "manifest.json",
        "SHA256SUMS",
    }
)
JSON_FILES = FILES - {"SHA256SUMS"}
SCHEMA_DIGEST = hashlib.sha256(b"butler.box5.a4.contracts.v3.1").hexdigest()
CODE_FILES = (
    "butler_pc_core/a4_verifier/cli.py",
    "butler_pc_core/a4_verifier/canonical.py",
    "butler_pc_core/a4_verifier/contracts.py",
    "butler_pc_core/a4_verifier/errors.py",
    "butler_pc_core/a4_verifier/receipt.py",
    "butler_pc_core/accounting/assignment/a4_store_schema_v32.py",
    "butler_pc_core/accounting/classify/reconciliation_v2.py",
    "butler_pc_core/accounting/classify/reconciliation_service_v2.py",
    "butler_pc_core/accounting/classify/source_snapshot_v2_1.py",
    "butler_pc_core/accounting/classify/verifier_authority.py",
    "butler_pc_core/accounting/classify/contracts/a4_v2/evidence_bundle.schema.json",
    "butler_pc_core/accounting/classify/contracts/a4_v31/code_dictionary.production.json",
    "butler_pc_core/accounting/classify/contracts/a4_v31/release_manifest.production.json",
    "butler_pc_core/accounting/classify/contracts/a4_v31/verifier_authority_trust.production.json",
    "butler_pc_core/accounting/classify/contracts/a4_v31/verifier_authority_trust.schema.json",
    "butler_pc_core/accounting/classify/contracts/a4_v31/verification_receipt.schema.json",
)
MAX_MONEY = 9_000_000_000_000_000
NAMESPACE = uuid.UUID("b8f2ceca-9183-5c4a-9750-3e0ed0f87066")
HEADER_KEYWORDS = frozenset(
    {
        "거래일",
        "거래일시",
        "거래일자",
        "일자",
        "일시",
        "거래내용",
        "적요",
        "내역",
        "비고",
        "메모",
        "출금",
        "출금액",
        "출금금액",
        "지급",
        "입금",
        "입금액",
        "입금금액",
        "금액",
        "거래금액",
        "잔액",
        "거래후잔액",
        "거래후 잔액",
        "거래처",
        "상대처",
        "받는분",
        "보내는분",
        "상대계좌예금주명",
    }
)


class Block(Exception):
    pass


def block(code: str) -> NoReturn:
    raise Block(code if re.fullmatch(r"BLOCK_[A-Z0-9_]+", code) else "BLOCK_VERIFY")


def pairs_hook(pairs: list[tuple[str, Any]]) -> dict[str, Any]:
    result: dict[str, Any] = {}
    for key, value in pairs:
        if key in result:
            block("BLOCK_DUPLICATE_KEY")
        result[key] = value
    return result


def strict_load(raw: bytes) -> Any:
    try:
        return json.loads(
            raw.decode("utf-8"),
            object_pairs_hook=pairs_hook,
            parse_constant=lambda _value: block("BLOCK_NUMBER"),
            parse_float=lambda _value: block("BLOCK_NUMBER"),
        )
    except Block:
        raise
    except Exception:
        block("BLOCK_SCHEMA")


def string_bytes(value: str) -> bytes:
    if any(0xD800 <= ord(char) <= 0xDFFF for char in value):
        block("BLOCK_SCHEMA")
    return json.dumps(value, ensure_ascii=False, separators=(",", ":")).encode("utf-8")


def jcs(value: Any) -> bytes:
    if value is None:
        return b"null"
    if value is True:
        return b"true"
    if value is False:
        return b"false"
    if type(value) is int:
        return str(value).encode("ascii")
    if isinstance(value, float):
        block("BLOCK_NUMBER")
    if isinstance(value, str):
        return string_bytes(value)
    if isinstance(value, list):
        return b"[" + b",".join(jcs(item) for item in value) + b"]"
    if isinstance(value, dict):
        if any(not isinstance(key, str) for key in value):
            block("BLOCK_SCHEMA")
        keys = sorted(value, key=lambda key: key.encode("utf-16-be", "surrogatepass"))
        return (
            b"{"
            + b",".join(string_bytes(key) + b":" + jcs(value[key]) for key in keys)
            + b"}"
        )
    block("BLOCK_SCHEMA")


def digest(value: Any) -> str:
    return hashlib.sha256(jcs(value)).hexdigest()


def read_files(root: Path) -> dict[str, bytes]:
    if root.is_symlink() or not root.is_dir():
        block("BLOCK_FILE_TYPE")
    observed: dict[str, bytes] = {}
    for item in root.iterdir():
        if item.name not in FILES:
            block("BLOCK_FILE_SET")
        mode = item.lstat().st_mode
        if (
            item.is_symlink()
            or not os.path.isfile(item)
            or not (mode & 0o170000) == 0o100000
        ):
            block("BLOCK_FILE_TYPE")
        observed[item.name] = item.read_bytes()
    if set(observed) != FILES:
        block("BLOCK_FILE_SET")
    return observed


def normalize_column(value: object) -> str:
    return re.sub(r"[\s　 ]+", "", re.sub(r"\(원\)", "", str(value)))


def normalized_cell(value: object) -> str:
    if value is None:
        return ""
    text = unicodedata.normalize("NFC", str(value)).strip()
    if len(text.encode("utf-8")) > 1_024:
        block("BLOCK_RESOURCE_LIMIT")
    return text


def read_source_frame(fd: int, suffix: str) -> Any:
    try:
        import pandas as pd
    except ImportError:
        block("BLOCK_VERIFIER_DEPENDENCY")

    def handle():
        duplicate = os.dup(fd)
        os.lseek(duplicate, 0, os.SEEK_SET)
        return os.fdopen(duplicate, "rb", closefd=True)

    if suffix == ".xlsx":
        try:
            import openpyxl
        except ImportError:
            block("BLOCK_VERIFIER_DEPENDENCY")
        with handle() as source:
            workbook = openpyxl.load_workbook(source, read_only=True, data_only=True)
            worksheet = workbook.active
            normalized_keywords = {normalize_column(item) for item in HEADER_KEYWORDS}
            best_row = 0
            best_score = 0.0
            for index, row in enumerate(
                worksheet.iter_rows(min_row=1, max_row=12, values_only=True)
            ):
                score = 0.0
                for cell in row:
                    if cell is None:
                        continue
                    text = normalize_column(str(cell).strip())
                    if text in normalized_keywords:
                        score += 1.0
                    elif any(keyword in text for keyword in normalized_keywords):
                        score += 0.5
                if score > best_score:
                    best_row, best_score = index, score
            workbook.close()
        with handle() as source:
            return pd.read_excel(source, header=best_row, dtype=str).fillna("")
    if suffix == ".xls":
        with handle() as source:
            return pd.read_excel(source, header=0, dtype=str).fillna("")
    if suffix != ".csv":
        block("BLOCK_FILE_TYPE")
    with handle() as source:
        raw = source.read()
    from io import BytesIO

    for encoding in ("utf-8-sig", "utf-8", "cp949", "euc-kr"):
        try:
            return pd.read_csv(BytesIO(raw), dtype=str, encoding=encoding).fillna("")
        except UnicodeDecodeError:
            continue
    block("BLOCK_INPUT_PARSE")


def rebuild_row_closure(frame: Any) -> dict[str, Any]:
    columns = tuple(str(column) for column in frame.columns)
    rows: list[dict[str, Any]] = []
    ordered = hashlib.sha256()
    for ordinal, (_, row) in enumerate(frame.iterrows()):
        cells = [[column, normalized_cell(row.get(column))] for column in columns]
        row_digest = digest(
            {"domain": "BUTLER_A4_PARSED_ROW_V2_1", "ordinal": ordinal, "cells": cells}
        )
        disposition = (
            "FEE"
            if normalized_cell(row.get("거래유형")).upper() in {"FEE", "수수료"}
            else "PRINCIPAL"
        )
        ordered.update(bytes.fromhex(row_digest))
        rows.append(
            {
                "ordinal": ordinal,
                "row_digest": row_digest,
                "disposition": disposition,
            }
        )
    return {
        "schema_version": "2.1.0",
        "ordered_row_root": ordered.hexdigest(),
        "row_count": len(rows),
        "disposition_counts": {
            name: sum(item["disposition"] == name for item in rows)
            for name in ("PRINCIPAL", "FEE", "HEADER", "QUARANTINED", "REJECTED")
        },
        "rows": rows,
    }


def physical_records(fd: int, suffix: str) -> list[list[object]]:
    duplicate = os.dup(fd)
    try:
        os.lseek(duplicate, 0, os.SEEK_SET)
        with os.fdopen(duplicate, "rb", closefd=True) as source:
            raw = source.read(512 * 1024 * 1024 + 1)
        duplicate = -1
    finally:
        if duplicate >= 0:
            os.close(duplicate)
    if len(raw) > 512 * 1024 * 1024:
        block("BLOCK_RESOURCE_LIMIT")
    if suffix == ".csv":
        text: str | None = None
        for encoding in ("utf-8-sig", "utf-8", "cp949", "euc-kr"):
            try:
                text = raw.decode(encoding)
                break
            except UnicodeDecodeError:
                continue
        if text is None:
            block("BLOCK_INPUT_CLOSURE")
        try:
            return [list(row) for row in csv.reader(io.StringIO(text, newline=""))]
        except csv.Error:
            block("BLOCK_INPUT_CLOSURE")
    if suffix == ".xlsx":
        workbook = None
        try:
            import openpyxl

            workbook = openpyxl.load_workbook(
                io.BytesIO(raw), read_only=True, data_only=False, keep_links=True
            )
            if len(workbook.worksheets) != 1:
                block("BLOCK_ADAPTER_POLICY")
            worksheet = workbook.worksheets[0]
            if worksheet.sheet_state != "visible" or getattr(
                workbook, "_external_links", ()
            ):
                block("BLOCK_ADAPTER_POLICY")
            rows: list[list[object]] = []
            for record in worksheet.iter_rows():
                if any(cell.data_type == "f" for cell in record):
                    block("BLOCK_ADAPTER_POLICY")
                rows.append([cell.value for cell in record])
            return rows
        except Block:
            raise
        except Exception:
            block("BLOCK_INPUT_CLOSURE")
        finally:
            if workbook is not None:
                workbook.close()
    if suffix == ".xls":
        try:
            import xlrd

            workbook = xlrd.open_workbook(file_contents=raw, on_demand=True)
            if workbook.nsheets != 1:
                block("BLOCK_ADAPTER_POLICY")
            sheet = workbook.sheet_by_index(0)
            if getattr(sheet, "visibility", 0) != 0:
                block("BLOCK_ADAPTER_POLICY")
            return [sheet.row_values(index) for index in range(sheet.nrows)]
        except Block:
            raise
        except Exception:
            block("BLOCK_INPUT_CLOSURE")
    block("BLOCK_UNSUPPORTED_SOURCE_FORMAT")


def rebuild_physical_row_closure(
    fd: int, suffix: str, source_sha256: str, adapter: Mapping[str, Any]
) -> dict[str, Any]:
    records = physical_records(fd, suffix)
    if not records or len(records) > 100_001:
        block("BLOCK_RESOURCE_LIMIT")
    required = tuple(
        str(adapter.get(name))
        for name in ("source_account", "bank_code", "booking_date", "withdrawal", "deposit")
    )
    normalized: list[list[str]] = []
    header_index: int | None = None
    for ordinal, record in enumerate(records):
        if len(record) > 256:
            block("BLOCK_RESOURCE_LIMIT")
        cells = [normalized_cell(value) for value in record]
        normalized.append(cells)
        if header_index is None and all(cells.count(name) == 1 for name in required):
            header_index = ordinal
    if header_index is None:
        block("BLOCK_ADAPTER_POLICY")
    headers = normalized[header_index]
    if len(headers) != len(set(headers)):
        block("BLOCK_AMBIGUOUS_COLUMN")
    fee_name = adapter.get("fee_type")
    fee_index = headers.index(str(fee_name)) if fee_name in headers else None
    rows: list[dict[str, Any]] = []
    ordered = hashlib.sha256()
    transaction_ordinal = 0
    for ordinal, cells in enumerate(normalized):
        if ordinal <= header_index:
            disposition, txn_ordinal = "HEADER", None
        elif not any(cells):
            disposition, txn_ordinal = "BLANK", None
        else:
            fee = (
                cells[fee_index].upper()
                if fee_index is not None and fee_index < len(cells)
                else ""
            )
            disposition = "FEE" if fee in {"FEE", "수수료"} else "PRINCIPAL"
            txn_ordinal = transaction_ordinal
            transaction_ordinal += 1
        locator = digest(
            {
                "domain": "BUTLER/A4/ROW_LOCATOR/V3.1",
                "source_sha256": source_sha256,
                "sheet_id": "sheet-0001",
                "row_ordinal": ordinal,
            }
        )
        cell_digest = digest(
            {
                "domain": "BUTLER/A4/CELL_VECTOR/V3.1",
                "row_ordinal": ordinal,
                "cells": cells,
            }
        )
        for value in (locator, cell_digest, disposition):
            encoded = value.encode("utf-8")
            ordered.update(len(encoded).to_bytes(4, "big"))
            ordered.update(encoded)
        rows.append(
            {
                "row_locator": locator,
                "sheet_id": "sheet-0001",
                "row_ordinal": ordinal,
                "disposition": disposition,
                "cell_vector_digest": cell_digest,
                "transaction_ordinal": txn_ordinal,
            }
        )
    names = ("PRINCIPAL", "FEE", "HEADER", "BLANK", "QUARANTINED", "REJECTED")
    return {
        "schema_version": "3.1.0",
        "ordered_row_root": ordered.hexdigest(),
        "row_count": len(rows),
        "transaction_row_count": transaction_ordinal,
        "disposition_counts": {
            name: sum(item["disposition"] == name for item in rows) for name in names
        },
        "rows": rows,
    }


def verify_source_snapshot(fd: int, suffix: str, receipt: Mapping[str, Any]) -> None:
    os.lseek(fd, 0, os.SEEK_SET)
    source_digest = hashlib.sha256()
    size = 0
    while True:
        chunk = os.read(fd, 1024 * 1024)
        if not chunk:
            break
        source_digest.update(chunk)
        size += len(chunk)
    os.lseek(fd, 0, os.SEEK_SET)
    snapshot = receipt.get("source_snapshot_receipt")
    required = {
        "schema_version",
        "source_file_sha256",
        "source_file_size",
        "pre_fstat_digest",
        "post_fstat_digest",
        "snapshot_sha256",
        "producer_verifier_fd_separated",
        "snapshot_path_unlinked",
    }
    if (
        not isinstance(snapshot, dict)
        or set(snapshot) != required
        or snapshot.get("schema_version") != "2.1.0"
        or snapshot.get("producer_verifier_fd_separated") is not True
        or snapshot.get("snapshot_path_unlinked") is not True
        or source_digest.hexdigest() != receipt.get("source_file_sha256")
        or source_digest.hexdigest() != snapshot.get("snapshot_sha256")
        or size != receipt.get("source_file_size")
        or size != snapshot.get("source_file_size")
    ):
        block("BLOCK_INPUT_CLOSURE")
    adapter = receipt.get("adapter_contract")
    if not isinstance(adapter, dict):
        block("BLOCK_SCHEMA")
    if (
        rebuild_physical_row_closure(
            fd, suffix, source_digest.hexdigest(), adapter
        )
        != receipt.get("row_closure")
    ):
        block("BLOCK_ROW_CLOSURE")


def _field(value: str) -> bytes:
    raw = value.encode("utf-8")
    return len(raw).to_bytes(4, "big") + raw


def _key_material(material: Any, tenant_id: str) -> dict[str, Any]:
    required = {
        "schema_version",
        "tenant_id",
        "key_id",
        "own_account_key_b64",
        "bank_reference_key_b64",
        "counterparty_account_key_b64",
        "run_transaction_key_b64",
    }
    if (
        not isinstance(material, dict)
        or set(material) != required
        or material.get("schema_version")
        != "butler.box5.a4.verifier_token_keys.v3.2"
        or material.get("tenant_id") != tenant_id
    ):
        block("BLOCK_TRUST_UNAVAILABLE")
    decoded_material = dict(material)
    for name in required - {"schema_version", "tenant_id", "key_id"}:
        try:
            decoded = base64.urlsafe_b64decode(str(decoded_material[name]))
        except Exception:
            block("BLOCK_TRUST_UNAVAILABLE")
        if len(decoded) != 32:
            block("BLOCK_TRUST_UNAVAILABLE")
        decoded_material[name] = decoded
    return decoded_material


def _dictionary(path: Path, expected_digest: str) -> dict[str, Any]:
    if path.is_symlink() or not path.is_file():
        block("BLOCK_DICTIONARY_VALUE")
    raw = path.read_bytes()
    if hashlib.sha256(raw).hexdigest() != expected_digest:
        block("BLOCK_DICTIONARY_VALUE")
    document = strict_load(raw)
    if (
        not isinstance(document, dict)
        or set(document)
        != {
            "schema_version", "dictionary_id", "version", "unknown_policy",
            "product_codes", "channels",
        }
        or document.get("schema_version") != "butler.box5.a4.code_dictionary.v3.1"
        or document.get("unknown_policy") != "QUARANTINE"
    ):
        block("BLOCK_DICTIONARY_VALUE")

    def aliases(items: Any) -> dict[str, str]:
        result: dict[str, str] = {}
        if not isinstance(items, list) or not items:
            block("BLOCK_DICTIONARY_VALUE")
        for item in items:
            if not isinstance(item, dict) or set(item) != {"id", "aliases"}:
                block("BLOCK_DICTIONARY_VALUE")
            stable_id = str(item["id"])
            if not re.fullmatch(r"[A-Za-z0-9._:-]{1,96}", stable_id):
                block("BLOCK_DICTIONARY_VALUE")
            for alias in item["aliases"]:
                key = unicodedata.normalize("NFC", str(alias)).strip().upper()
                if not key or key in result:
                    block("BLOCK_DICTIONARY_VALUE")
                result[key] = stable_id
        if "DEFAULT" not in result.values():
            block("BLOCK_DICTIONARY_VALUE")
        return result

    return {
        "product": aliases(document["product_codes"]),
        "channel": aliases(document["channels"]),
    }


def _account_number(value: Any) -> str:
    return re.sub(r"[^0-9A-Za-z]", "", normalized_cell(value).upper())


def _token(key: bytes, *values: str) -> str:
    return hmac.new(key, b"".join(_field(value) for value in values), hashlib.sha256).hexdigest()


def _source_amount(value: Any) -> int:
    text = normalized_cell(value)
    if not re.fullmatch(r"(?:0|[1-9][0-9]{0,18}|[1-9][0-9]{0,2}(?:,[0-9]{3})+)", text):
        block("BLOCK_SOURCE_SEMANTIC_MISMATCH")
    amount = int(text.replace(",", ""))
    if not 1 <= amount <= MAX_MONEY:
        block("BLOCK_RANGE")
    return amount


def independently_compile_source(
    frame: Any,
    input_receipt: Mapping[str, Any],
    run: Mapping[str, Any],
    accounts: Sequence[Mapping[str, Any]],
    material: Mapping[str, Any],
    dictionary: Mapping[str, Mapping[str, str]],
) -> list[dict[str, Any]]:
    """Compile raw rows without importing or calling producer code."""

    adapter = input_receipt.get("adapter_contract")
    if not isinstance(adapter, dict):
        block("BLOCK_SOURCE_SEMANTIC_MISMATCH")
    columns = {unicodedata.normalize("NFC", str(value)).strip(): value for value in frame.columns}
    names = {
        key: adapter.get(key)
        for key in (
            "source_account", "bank_code", "booking_date", "withdrawal", "deposit",
            "reference", "counter_account", "fee_type", "product_code", "channel",
            "currency", "value_date", "duplicate_of_reference",
            "reversal_of_reference",
        )
    }
    for key in ("source_account", "bank_code", "booking_date", "withdrawal", "deposit"):
        if names[key] not in columns:
            block("BLOCK_SOURCE_SEMANTIC_MISMATCH")
    registry: dict[tuple[str, str, str], str] = {}
    for account in accounts:
        try:
            key = (
                str(account["bank_code"]),
                str(account["lookup_key_id"]),
                str(account["lookup_token"]),
            )
            registry[key] = uuid_string(account["account_id"])
        except Exception:
            block("BLOCK_REGISTRY_IDENTITY")
    rows = tuple(frame.iterrows())
    identities: dict[tuple[str, str], tuple[str, str]] = {}
    number_ids: dict[str, set[str]] = {}
    tenant = str(run["tenant_id"])
    for _, row in rows:
        bank = normalized_cell(row.get(columns[names["bank_code"]])).upper()
        number = _account_number(row.get(columns[names["source_account"]]))
        account_token = _token(material["own_account_key_b64"], tenant, bank, number)
        account_id = registry.get((bank, str(material["key_id"]), account_token))
        if not bank or not number or account_id is None:
            block("BLOCK_REGISTRY_IDENTITY")
        identities[(bank, number)] = (account_id, account_token)
        number_ids.setdefault(number, set()).add(account_id)
    namespace = uuid.uuid5(uuid.UUID(tenant), str(input_receipt["source_file_sha256"]))
    rebuilt: list[dict[str, Any]] = []
    artifacts: set[str] = set()
    formats = adapter.get("allowed_date_formats")
    if not isinstance(formats, list):
        block("BLOCK_SOURCE_SEMANTIC_MISMATCH")
    for ordinal, (_, row) in enumerate(rows):
        withdrawal = normalized_cell(row.get(columns[names["withdrawal"]]))
        deposit = normalized_cell(row.get(columns[names["deposit"]]))
        if bool(withdrawal) == bool(deposit):
            block("BLOCK_SOURCE_SEMANTIC_MISMATCH")
        direction = "OUTFLOW" if withdrawal else "INFLOW"
        amount = _source_amount(withdrawal or deposit)
        bank = normalized_cell(row.get(columns[names["bank_code"]])).upper()
        number = _account_number(row.get(columns[names["source_account"]]))
        identity = identities.get((bank, number))
        if identity is None:
            block("BLOCK_REGISTRY_IDENTITY")
        account_id = identity[0]
        parsed = None
        raw_date = normalized_cell(row.get(columns[names["booking_date"]]))
        for date_format in formats:
            try:
                parsed = datetime.strptime(raw_date, str(date_format))
                break
            except ValueError:
                continue
        if parsed is None:
            block("BLOCK_SOURCE_SEMANTIC_MISMATCH")
        local_date = parsed.date().isoformat()
        booked_at = parsed.replace(tzinfo=ZoneInfo(str(adapter["zone_name"]))).astimezone(timezone.utc)
        booked_at_text = booked_at.isoformat().replace("+00:00", "Z")
        reference_token = None
        if names["reference"] in columns:
            reference = normalized_cell(row.get(columns[names["reference"]]))
            if reference:
                reference_token = _token(
                    material["bank_reference_key_b64"], tenant, "BANK_REFERENCE", reference
                )
        counter_token = None
        resolved = None
        resolution = "MISSING"
        if names["counter_account"] in columns:
            raw_counter = normalized_cell(row.get(columns[names["counter_account"]]))
            counter = _account_number(raw_counter)
            if any(marker in raw_counter for marker in ("*", "•", "●")):
                resolution = "MASKED"
            elif counter:
                counter_token = _token(
                    material["counterparty_account_key_b64"],
                    tenant,
                    "COUNTERPARTY_ACCOUNT",
                    counter,
                )
                candidates = number_ids.get(counter, set())
                if len(candidates) == 1:
                    resolved, resolution = next(iter(candidates)), "EXACT_VERIFIED"
                elif len(candidates) > 1:
                    block("BLOCK_COUNTERPARTY_CONFLICT")
                else:
                    resolution = "UNRESOLVED"
        row_kind = "PRINCIPAL"
        if names["fee_type"] in columns:
            fee = normalized_cell(row.get(columns[names["fee_type"]])).upper()
            if fee:
                if fee not in {"FEE", "수수료"}:
                    block("BLOCK_SOURCE_SEMANTIC_MISMATCH")
                row_kind = "FEE"
        product_value = "DEFAULT"
        if names["product_code"] in columns:
            product_value = normalized_cell(row.get(columns[names["product_code"]])) or "DEFAULT"
        channel_value = "DEFAULT"
        if names["channel"] in columns:
            channel_value = normalized_cell(row.get(columns[names["channel"]])) or "DEFAULT"
        product_id = dictionary["product"].get(product_value.upper())
        channel_id = dictionary["channel"].get(channel_value.upper())
        if product_id is None or channel_id is None:
            block("BLOCK_DICTIONARY_VALUE")
        currency = "KRW"
        if names["currency"] in columns:
            currency = normalized_cell(row.get(columns[names["currency"]])).upper()
            if not re.fullmatch(r"[A-Z]{3}", currency):
                block("BLOCK_SOURCE_SEMANTIC_MISMATCH")
        value_date = None
        if names["value_date"] in columns:
            raw_value_date = normalized_cell(row.get(columns[names["value_date"]]))
            if raw_value_date:
                parsed_value_date = None
                for date_format in formats:
                    try:
                        parsed_value_date = datetime.strptime(
                            raw_value_date, str(date_format)
                        )
                        break
                    except ValueError:
                        continue
                if parsed_value_date is None:
                    block("BLOCK_SOURCE_SEMANTIC_MISMATCH")
                value_date = parsed_value_date.date().isoformat()
        duplicate_of_reference_token = None
        reversal_of_reference_token = None
        for name, destination in (
            ("duplicate_of_reference", "duplicate"),
            ("reversal_of_reference", "reversal"),
        ):
            if names[name] not in columns:
                continue
            raw_link = normalized_cell(row.get(columns[names[name]]))
            if not raw_link:
                continue
            link_token = _token(
                material["bank_reference_key_b64"],
                tenant,
                "BANK_REFERENCE",
                raw_link,
            )
            if destination == "duplicate":
                duplicate_of_reference_token = link_token
            else:
                reversal_of_reference_token = link_token
        if (
            duplicate_of_reference_token is not None
            and reversal_of_reference_token is not None
        ):
            block("BLOCK_REVERSAL_BINDING")
        canonical = {
            "account_id": account_id,
            "direction": direction,
            "amount_minor": amount,
            "currency": currency,
            "booked_at_utc": booked_at_text,
            "local_date": local_date,
            "bank_code": bank,
            "bank_reference_token": reference_token,
            "counterparty_account_token": counter_token,
            "resolved_counterparty_account_id": resolved,
            "counterparty_resolution": resolution,
            "row_kind": row_kind,
            "product_code_id": product_id,
            "channel_id": channel_id,
            "adapter_id": adapter["adapter_id"],
            "adapter_digest": run["adapter_digest"],
            "registry_digest": run["registry_digest"],
            "value_date": value_date,
            "duplicate_of_reference_token": duplicate_of_reference_token,
            "reversal_of_reference_token": reversal_of_reference_token,
        }
        canonical_digest = digest(canonical)
        source_row_digest = digest(
            {
                "source_file_sha256": input_receipt["source_file_sha256"],
                "row_ordinal": ordinal,
                "canonical_row_sha256": canonical_digest,
            }
        )
        txn_uid = str(uuid.uuid5(
            namespace,
            f"{input_receipt['source_file_sha256']}:{ordinal}:{canonical_digest}",
        ))
        artifact = _token(
            material["run_transaction_key_b64"],
            tenant,
            "RUN_TRANSACTION",
            str(run["run_id"]) + tenant + txn_uid,
        )[:32]
        if artifact in artifacts:
            block("BLOCK_ARTIFACT_TOKEN_COLLISION")
        artifacts.add(artifact)
        rebuilt.append(
            {
                "tenant_id": tenant,
                "txn_uid": txn_uid,
                "source_row_digest": source_row_digest,
                "row_ordinal": ordinal,
                "canonical_row_digest": canonical_digest,
                "account_id": account_id,
                "direction": direction,
                "amount_minor": amount,
                "currency": currency,
                "booked_at_utc": booked_at_text,
                "local_date": local_date,
                "artifact_token": artifact,
                "bank_code": bank,
                "bank_reference_token": reference_token,
                "counterparty_account_token": counter_token,
                "resolved_counterparty_account_id": resolved,
                "counterparty_resolution": resolution,
                "row_kind": row_kind,
                "product_code_id": product_id,
                "channel_id": channel_id,
                "adapter_id": adapter["adapter_id"],
                "adapter_digest": run["adapter_digest"],
                "registry_digest": run["registry_digest"],
                "value_date": value_date,
                "duplicate_of_reference_token": duplicate_of_reference_token,
                "reversal_of_reference_token": reversal_of_reference_token,
            }
        )
    return sorted(rebuilt, key=lambda item: (item["tenant_id"], item["txn_uid"]))


def verify_hashes(raw: Mapping[str, bytes], docs: Mapping[str, Any]) -> None:
    expected_lines = b"".join(
        f"{hashlib.sha256(raw[name]).hexdigest()}  {name}\n".encode("ascii")
        for name in sorted(JSON_FILES)
    )
    if raw["SHA256SUMS"] != expected_lines:
        block("BLOCK_DIGEST")
    manifest = docs["manifest.json"]
    if not isinstance(manifest, dict) or set(manifest) != {
        "schema_version",
        "run_id",
        "nonce_digest",
        "code_tree_oid",
        "files",
    }:
        block("BLOCK_SCHEMA")
    files = manifest["files"]
    if not isinstance(files, list) or len(files) != 8:
        block("BLOCK_FILE_SET")
    observed: dict[str, tuple[str, int]] = {}
    for item in files:
        if not isinstance(item, dict) or set(item) != {"path", "sha256", "size"}:
            block("BLOCK_SCHEMA")
        path = item["path"]
        if path in observed or path not in JSON_FILES - {"manifest.json"}:
            block("BLOCK_FILE_SET")
        observed[path] = (item["sha256"], item["size"])
    if set(observed) != JSON_FILES - {"manifest.json"}:
        block("BLOCK_FILE_SET")
    for name, (sha, size) in observed.items():
        if (
            type(size) is not int
            or size != len(raw[name])
            or sha != hashlib.sha256(raw[name]).hexdigest()
        ):
            block("BLOCK_DIGEST")


def parse_time(value: Any) -> datetime:
    if not isinstance(value, str) or not value.endswith("Z"):
        block("BLOCK_TIMEZONE")
    try:
        result = datetime.fromisoformat(value[:-1] + "+00:00")
    except ValueError:
        block("BLOCK_TIMEZONE")
    if result.tzinfo is None or result.utcoffset() != timezone.utc.utcoffset(result):
        block("BLOCK_TIMEZONE")
    return result


def load_trust(path: Path) -> dict[str, Any]:
    if path.is_symlink() or not path.is_file():
        block("BLOCK_UNPINNED_KEY")
    trust = strict_load(path.read_bytes())
    required = {
        "schema_version",
        "policy_type",
        "algorithm",
        "expected_signer_key_id",
        "public_key_b64",
        "revocation_epoch",
        "enabled",
    }
    if not isinstance(trust, dict) or set(trust) != required:
        block("BLOCK_UNPINNED_KEY")
    if (
        trust["schema_version"] != "1.0.0"
        or trust["enabled"] is not True
        or trust["policy_type"] != "A4_RECON_POLICY"
        or trust["algorithm"] != "Ed25519"
    ):
        block("BLOCK_UNPINNED_KEY")
    if type(trust["revocation_epoch"]) is not int or trust["revocation_epoch"] < 0:
        block("BLOCK_UNPINNED_KEY")
    return trust


def verify_policy(
    receipt: Any, payload: Any, trust: Mapping[str, Any], expected: Mapping[str, str]
) -> dict[str, Any]:
    required = {
        "policy_type",
        "policy_id",
        "policy_version",
        "policy_digest",
        "finance_approval_id",
        "finance_approver_role",
        "signer_key_id",
        "issued_at",
        "effective_from",
        "expires_at",
        "revocation_epoch",
        "tenant_id",
        "schema_digest",
        "adapter_digest",
        "account_snapshot_digest",
        "dictionary_digest",
        "registry_digest",
        "code_tree_oid",
        "run_nonce_digest",
        "signature_b64",
    }
    if (
        not isinstance(receipt, dict)
        or set(receipt) != required
        or not isinstance(payload, dict)
    ):
        block("BLOCK_POLICY")
    if (
        receipt["policy_type"] != trust["policy_type"]
        or receipt["signer_key_id"] != trust["expected_signer_key_id"]
    ):
        block("BLOCK_SIGNER")
    if receipt["finance_approver_role"] != "FINANCE_APPROVER":
        block("BLOCK_SIGNER")
    if receipt["revocation_epoch"] != trust["revocation_epoch"]:
        block("BLOCK_REVOKED")
    for key in (
        "tenant_id",
        "schema_digest",
        "adapter_digest",
        "account_snapshot_digest",
        "dictionary_digest",
        "registry_digest",
        "code_tree_oid",
        "run_nonce_digest",
    ):
        if receipt[key] != expected[key]:
            block("BLOCK_BINDING")
    now = datetime.now(timezone.utc)
    issued, start, expires = (
        parse_time(receipt[key])
        for key in ("issued_at", "effective_from", "expires_at")
    )
    if not issued <= now <= expires or now < start or expires <= start:
        block("BLOCK_EXPIRED")
    if receipt["policy_digest"] != digest(payload):
        block("BLOCK_DIGEST")
    unsigned = {key: value for key, value in receipt.items() if key != "signature_b64"}
    try:
        public = base64.b64decode(trust["public_key_b64"], validate=True)
        signature = base64.b64decode(receipt["signature_b64"], validate=True)
        Ed25519PublicKey.from_public_bytes(public).verify(
            signature, jcs({"receipt": unsigned, "payload": payload})
        )
    except (ValueError, TypeError, InvalidSignature):
        block("BLOCK_SIGNATURE")
    return payload


def uuid_string(value: Any) -> str:
    if not isinstance(value, str):
        block("BLOCK_SCHEMA")
    try:
        parsed = str(uuid.UUID(value))
    except ValueError:
        block("BLOCK_SCHEMA")
    if parsed != value.lower():
        block("BLOCK_SCHEMA")
    return parsed


def verify_bindings(
    docs: Mapping[str, Any],
) -> tuple[dict[str, Any], list[dict[str, Any]], dict[str, Any]]:
    run = docs["run_request.json"]
    input_receipt = docs["input_receipt.json"]
    account = docs["account_snapshot_receipt.json"]
    if not all(isinstance(item, dict) for item in (run, input_receipt, account)):
        block("BLOCK_SCHEMA")
    required_run = {
        "schema_version",
        "run_id",
        "nonce",
        "tenant_id",
        "code_tree_oid",
        "input_closure_digest",
        "adapter_digest",
        "policy_digest",
        "account_snapshot_digest",
        "dictionary_digest",
        "registry_digest",
    }
    if set(run) != required_run or run["schema_version"] != "1.0.0":
        block("BLOCK_SCHEMA")
    run_id, tenant = uuid_string(run["run_id"]), uuid_string(run["tenant_id"])
    if not re.fullmatch(r"[0-9a-fA-F]{64}", run["nonce"]):
        block("BLOCK_RUN_BINDING")
    nonce_digest = hashlib.sha256(bytes.fromhex(run["nonce"])).hexdigest()
    for name in (
        "input_receipt.json",
        "account_snapshot_receipt.json",
        "graph_observation.json",
        "classification_receipt.json",
        "runtime_invariants.json",
        "dlp_receipt.json",
        "manifest.json",
    ):
        item = docs[name]
        if item.get("run_id") != run_id:
            block("BLOCK_RUN_BINDING")
        if "nonce_digest" in item and item["nonce_digest"] != nonce_digest:
            block("BLOCK_RUN_BINDING")
    if input_receipt.get("tenant_id") != tenant or account.get("tenant_id") != tenant:
        block("BLOCK_TENANT")
    source = {
        "source_file_sha256": input_receipt.get("source_file_sha256"),
        "source_file_size": input_receipt.get("source_file_size"),
    }
    if (
        digest(source) != run["input_closure_digest"]
        or input_receipt.get("input_closure_digest") != run["input_closure_digest"]
    ):
        block("BLOCK_INPUT_CLOSURE")
    adapter = input_receipt.get("adapter_contract")
    if (
        not isinstance(adapter, dict)
        or digest(adapter) != run["adapter_digest"]
        or input_receipt.get("adapter_digest") != run["adapter_digest"]
        or input_receipt.get("dictionary_digest") != run["dictionary_digest"]
        or input_receipt.get("registry_digest") != run["registry_digest"]
    ):
        block("BLOCK_BINDING")
    accounts = account.get("accounts")
    if (
        not isinstance(accounts, list)
        or digest(accounts) != run["account_snapshot_digest"]
        or account.get("account_snapshot_digest") != run["account_snapshot_digest"]
        or account.get("registry_digest") != run["registry_digest"]
    ):
        block("BLOCK_BINDING")
    txs = input_receipt.get("compiled_transactions")
    row_closure = input_receipt.get("row_closure")
    if (
        not isinstance(txs, list)
        or not isinstance(row_closure, dict)
        or input_receipt.get("source_row_count") != len(txs)
        or row_closure.get("transaction_row_count") != len(txs)
        or not isinstance(row_closure.get("rows"), list)
        or len(row_closure["rows"]) != row_closure.get("row_count")
    ):
        block("BLOCK_SCHEMA")
    namespace = uuid.uuid5(uuid.UUID(tenant), input_receipt["source_file_sha256"])
    uids: dict[str, str] = {}
    rows: dict[str, str] = {}
    artifacts: set[str] = set()
    account_ids = {
        item.get("account_id") for item in accounts if isinstance(item, dict)
    }
    for tx in txs:
        required_tx = {
            "tenant_id",
            "txn_uid",
            "source_row_digest",
            "row_ordinal",
            "canonical_row_digest",
            "account_id",
            "direction",
            "amount_minor",
            "currency",
            "booked_at_utc",
            "local_date",
            "artifact_token",
            "bank_code",
            "bank_reference_token",
            "counterparty_account_token",
            "resolved_counterparty_account_id",
            "counterparty_resolution",
            "row_kind",
            "product_code_id",
            "channel_id",
            "adapter_id",
            "adapter_digest",
            "registry_digest",
            "value_date",
            "duplicate_of_reference_token",
            "reversal_of_reference_token",
        }
        if not isinstance(tx, dict) or set(tx) != required_tx:
            block("BLOCK_SCHEMA")
        uid = uuid_string(tx.get("txn_uid"))
        if tx.get("tenant_id") != tenant or tx.get("account_id") not in account_ids:
            block("BLOCK_TENANT")
        ordinal, canonical = tx.get("row_ordinal"), tx.get("canonical_row_digest")
        if (
            type(ordinal) is not int
            or ordinal < 0
            or not re.fullmatch(r"[0-9a-f]{64}", str(canonical))
        ):
            block("BLOCK_SCHEMA")
        row_digest = digest(
            {
                "source_file_sha256": input_receipt["source_file_sha256"],
                "row_ordinal": ordinal,
                "canonical_row_sha256": canonical,
            }
        )
        if tx.get("source_row_digest") != row_digest:
            block("BLOCK_DIGEST")
        expected_uid = str(
            uuid.uuid5(
                namespace,
                f"{input_receipt['source_file_sha256']}:{ordinal}:{canonical}",
            )
        )
        if uid != expected_uid:
            block("BLOCK_TXN_UID_CONFLICT")
        if uid in uids and uids[uid] != row_digest:
            block("BLOCK_TXN_UID_CONFLICT")
        if row_digest in rows and rows[row_digest] != uid:
            block("BLOCK_DUPLICATE_SOURCE_ROW")
        token = tx.get("artifact_token")
        if (
            not isinstance(token, str)
            or not re.fullmatch(r"[0-9a-f]{32}", token)
            or token in artifacts
        ):
            block("BLOCK_ARTIFACT_TOKEN_COLLISION")
        amount = tx.get("amount_minor")
        if type(amount) is not int or not 1 <= amount <= MAX_MONEY:
            block("BLOCK_RANGE")
        if not re.fullmatch(r"[A-Z]{3}", str(tx.get("currency"))) or tx.get("direction") not in {
            "INFLOW",
            "OUTFLOW",
        }:
            block("BLOCK_SCHEMA")
        if (
            tx.get("adapter_id") != adapter.get("adapter_id")
            or tx.get("adapter_digest") != run["adapter_digest"]
            or tx.get("registry_digest") != run["registry_digest"]
            or not re.fullmatch(r"[A-Za-z0-9._:-]{1,96}", str(tx.get("product_code_id")))
            or not re.fullmatch(r"[A-Za-z0-9._:-]{1,96}", str(tx.get("channel_id")))
        ):
            block("BLOCK_BINDING")
        value_date = tx.get("value_date")
        if value_date is not None:
            try:
                if datetime.fromisoformat(str(value_date)).date().isoformat() != value_date:
                    block("BLOCK_SCHEMA")
            except (TypeError, ValueError):
                block("BLOCK_SCHEMA")
        duplicate_link = tx.get("duplicate_of_reference_token")
        reversal_link = tx.get("reversal_of_reference_token")
        if (
            any(
                link is not None
                and not re.fullmatch(r"[0-9a-f]{64}", str(link))
                for link in (duplicate_link, reversal_link)
            )
            or (duplicate_link is not None and reversal_link is not None)
        ):
            block("BLOCK_SCHEMA")
        resolution = tx.get("counterparty_resolution")
        resolved = tx.get("resolved_counterparty_account_id")
        if resolution not in {
            "EXACT_VERIFIED",
            "MASKED",
            "MISSING",
            "UNRESOLVED",
            "CONFLICT",
        }:
            block("BLOCK_SCHEMA")
        if resolution == "CONFLICT":
            block("BLOCK_COUNTERPARTY_CONFLICT")
        if resolution == "EXACT_VERIFIED":
            if uuid_string(resolved) not in account_ids:
                block("BLOCK_ACCOUNT_IDENTITY")
        elif resolved is not None:
            block("BLOCK_ACCOUNT_IDENTITY")
        parse_time(tx.get("booked_at_utc"))
        uids[uid], rows[row_digest] = row_digest, uid
        artifacts.add(token)
    expected = {
        "tenant_id": tenant,
        "schema_digest": SCHEMA_DIGEST,
        "adapter_digest": run["adapter_digest"],
        "account_snapshot_digest": run["account_snapshot_digest"],
        "dictionary_digest": run["dictionary_digest"],
        "registry_digest": run["registry_digest"],
        "code_tree_oid": run["code_tree_oid"],
        "run_nonce_digest": nonce_digest,
    }
    return run, txs, expected


def exact_ref(left: Mapping[str, Any], right: Mapping[str, Any]) -> bool:
    return bool(left.get("bank_reference_token")) and left.get(
        "bank_reference_token"
    ) == right.get("bank_reference_token")


def edge_digest(edge: Mapping[str, Any]) -> str:
    payload = {
        "outflow": edge["outflow_txn_uid"],
        "inflow": edge["inflow_txn_uid"],
        "fee": edge["fee_txn_uid"],
        "mode": edge["fee_mode"],
        "fee_minor": edge["fee_minor"],
        "eligibility_basis": edge["eligibility_basis"],
        "binding_digest": edge["binding_digest"],
    }
    if edge.get("outflow_txn_uids") is not None or edge.get("match_type") == "FX":
        payload.update(
            {
                "outflows": sorted(edge["outflow_txn_uids"]),
                "inflows": sorted(edge["inflow_txn_uids"]),
                "match_type": edge["match_type"],
                "currency_mode": edge["currency_mode"],
                "fx_rule_id": edge["fx_rule_id"],
                "fx_receipt_digest": edge["fx_receipt_digest"],
            }
        )
    return digest(payload)


def strong_pair(
    outflow: Mapping[str, Any],
    inflow: Mapping[str, Any],
    policy: Mapping[str, Any],
) -> bool:
    left_day = datetime.fromisoformat(
        outflow.get("value_date") or outflow["local_date"]
    ).date()
    right_day = datetime.fromisoformat(
        inflow.get("value_date") or inflow["local_date"]
    ).date()
    gap = (right_day - left_day).days
    return (
        outflow["direction"] == "OUTFLOW"
        and inflow["direction"] == "INFLOW"
        and outflow["tenant_id"] == inflow["tenant_id"]
        and outflow["account_id"] != inflow["account_id"]
        and outflow["counterparty_resolution"] == "EXACT_VERIFIED"
        and inflow["counterparty_resolution"] == "EXACT_VERIFIED"
        and outflow["resolved_counterparty_account_id"] == inflow["account_id"]
        and inflow["resolved_counterparty_account_id"] == outflow["account_id"]
        and policy["min_signed_gap_days"] <= gap <= policy["max_signed_gap_days"]
    )


def terminal_partition(
    txs: Sequence[dict[str, Any]],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """Independently enforce duplicate/reversal terminal ownership."""

    reference_index: dict[tuple[str, str], list[dict[str, Any]]] = {}
    for tx in txs:
        reference = tx.get("bank_reference_token")
        if reference is not None:
            reference_index.setdefault(
                (str(tx["tenant_id"]), str(reference)), []
            ).append(tx)

    claimed: set[str] = set()
    groups: list[dict[str, Any]] = []
    for tx in txs:
        link = tx.get("reversal_of_reference_token")
        if link is None:
            continue
        targets = reference_index.get((str(tx["tenant_id"]), str(link)), [])
        if len(targets) != 1:
            block("BLOCK_REVERSAL_BINDING")
        target = targets[0]
        uids = sorted((str(tx["txn_uid"]), str(target["txn_uid"])))
        if (
            uids[0] == uids[1]
            or tx["row_kind"] != "PRINCIPAL"
            or target["row_kind"] != "PRINCIPAL"
            or tx["account_id"] != target["account_id"]
            or tx["direction"] == target["direction"]
            or tx["amount_minor"] != target["amount_minor"]
            or tx["currency"] != target["currency"]
            or claimed.intersection(uids)
        ):
            block("BLOCK_REVERSAL_BINDING")
        claimed.update(uids)
        reason = "BLOCK_REVERSAL_BINDING"
        groups.append(
            {
                "group_id": digest(
                    {"transaction_uids": uids, "reason_code": reason}
                ),
                "state": "BLOCKED",
                "reason_code": reason,
                "transaction_uids": uids,
            }
        )

    for tx in txs:
        link = tx.get("duplicate_of_reference_token")
        if link is None:
            continue
        targets = reference_index.get((str(tx["tenant_id"]), str(link)), [])
        uid = str(tx["txn_uid"])
        if (
            len(targets) != 1
            or targets[0]["txn_uid"] == uid
            or uid in claimed
        ):
            block("BLOCK_DUPLICATE_BINDING")
        claimed.add(uid)
        reason = "BLOCK_DUPLICATE_BINDING"
        uids = [uid]
        groups.append(
            {
                "group_id": digest(
                    {"transaction_uids": uids, "reason_code": reason}
                ),
                "state": "BLOCKED",
                "reason_code": reason,
                "transaction_uids": uids,
            }
        )

    active = [tx for tx in txs if str(tx["txn_uid"]) not in claimed]
    return active, sorted(groups, key=lambda item: item["group_id"])


def multi_binding(
    outflows: Sequence[Mapping[str, Any]],
    inflows: Sequence[Mapping[str, Any]],
    match_type: str,
    fx_rule: Mapping[str, Any] | None = None,
) -> str:
    return digest(
        {
            "domain": "BUTLER_A4_STRONG_BINDING_V5_4",
            "tenant_id": outflows[0]["tenant_id"],
            "outflow_txn_uids": sorted(item["txn_uid"] for item in outflows),
            "inflow_txn_uids": sorted(item["txn_uid"] for item in inflows),
            "outflow_account_ids": sorted({item["account_id"] for item in outflows}),
            "inflow_account_ids": sorted({item["account_id"] for item in inflows}),
            "match_type": match_type,
            "fx_rule_id": fx_rule.get("rule_id") if fx_rule else None,
            "fx_receipt_digest": (
                fx_rule.get("rate_source_receipt_digest") if fx_rule else None
            ),
            "eligibility_basis": "RECIPROCAL_ACCOUNT",
        }
    )


def aggregate_delta_ms(items: Sequence[Mapping[str, Any]]) -> int:
    times = [parse_time(item["booked_at_utc"]) for item in items]
    return int((max(times) - min(times)).total_seconds() * 1_000)


def exact_fx_rule(
    rule: Mapping[str, Any],
    outflow: Mapping[str, Any],
    inflow: Mapping[str, Any],
) -> bool:
    required = {
        "rule_id", "from_currency", "to_currency", "rate_numerator",
        "rate_denominator", "source_fee_minor", "target_fee_minor",
        "spread_bps", "rounding_mode", "effective_from", "effective_to",
        "rate_source_receipt_digest",
    }
    if not isinstance(rule, dict) or set(rule) != required:
        block("BLOCK_POLICY")
    integers = [
        rule["rate_numerator"], rule["rate_denominator"],
        rule["source_fee_minor"], rule["target_fee_minor"], rule["spread_bps"],
    ]
    if any(type(value) is not int for value in integers) or not (
        1 <= integers[0] <= 1_000_000_000_000
        and 1 <= integers[1] <= 1_000_000_000_000
        and 0 <= integers[2] <= MAX_MONEY
        and 0 <= integers[3] <= MAX_MONEY
        and 0 <= integers[4] < 10_000
        and rule["rounding_mode"] == "EXACT_RATIONAL"
        and re.fullmatch(r"[A-Z]{3}", str(rule["from_currency"]))
        and re.fullmatch(r"[A-Z]{3}", str(rule["to_currency"]))
        and rule["from_currency"] != rule["to_currency"]
        and re.fullmatch(r"[A-Za-z0-9._:-]{1,96}", str(rule["rule_id"]))
        and re.fullmatch(r"[0-9a-f]{64}", str(rule["rate_source_receipt_digest"]))
    ):
        block("BLOCK_POLICY")
    start, end = parse_time(rule["effective_from"]), parse_time(rule["effective_to"])
    source_time, target_time = parse_time(outflow["booked_at_utc"]), parse_time(inflow["booked_at_utc"])
    source = outflow["amount_minor"] - rule["source_fee_minor"]
    target = inflow["amount_minor"] + rule["target_fee_minor"]
    return (
        rule["from_currency"] == outflow["currency"]
        and rule["to_currency"] == inflow["currency"]
        and start <= source_time <= end
        and start <= target_time <= end
        and source > 0
        and target > 0
        and source * rule["rate_numerator"] * (10_000 - rule["spread_bps"])
        == target * rule["rate_denominator"] * 10_000
    )


def rebuild_edges(
    txs: Sequence[dict[str, Any]], policy: Mapping[str, Any]
) -> tuple[list[dict[str, Any]], int, int, int]:
    required_policy = {
        "min_signed_gap_days",
        "max_signed_gap_days",
        "max_component_nodes",
        "max_candidate_edges",
        "max_transactions",
        "max_pair_checks",
        "max_trace_bucket",
        "max_memory_bytes",
        "fee_rules",
    }
    optional_policy = {"max_subset_size", "max_subset_branches", "fx_rules"}
    if (
        not required_policy <= set(policy) <= required_policy | optional_policy
        or not isinstance(policy["fee_rules"], list)
        or not isinstance(policy.get("fx_rules", []), list)
    ):
        block("BLOCK_POLICY")
    subset_cap = policy.get("max_subset_size", 8)
    branch_cap = policy.get("max_subset_branches", 4_096)
    if (
        type(subset_cap) is not int
        or not 2 <= subset_cap <= 8
        or type(branch_cap) is not int
        or not 1 <= branch_cap <= 65_536
    ):
        block("BLOCK_POLICY")
    caps = (
        policy["max_transactions"],
        policy["max_pair_checks"],
        policy["max_trace_bucket"],
        policy["max_memory_bytes"],
        policy["max_candidate_edges"],
    )
    if any(type(value) is not int or value <= 0 for value in caps):
        block("BLOCK_POLICY")
    if len(txs) > policy["max_transactions"]:
        block("BLOCK_RESOURCE_TRANSACTION_COUNT")
    if len(txs) * 2_048 > policy["max_memory_bytes"]:
        block("BLOCK_RESOURCE_MEMORY")
    principals = sorted(
        (tx for tx in txs if tx.get("row_kind") == "PRINCIPAL"),
        key=lambda tx: tx["txn_uid"],
    )
    fees = [tx for tx in txs if tx.get("row_kind") == "FEE"]
    inflow_index: dict[tuple[str, str, str, int], list[dict[str, Any]]] = {}
    for inflow in principals:
        if inflow["direction"] == "INFLOW":
            key = (
                inflow["tenant_id"],
                inflow["account_id"],
                inflow["currency"],
                inflow["amount_minor"],
            )
            inflow_index.setdefault(key, []).append(inflow)
    for bucket in inflow_index.values():
        bucket.sort(key=lambda tx: (tx["booked_at_utc"], tx["txn_uid"]))
    fee_index: dict[tuple[str, str, str, str], list[dict[str, Any]]] = {}
    for fee in fees:
        reference = fee.get("bank_reference_token")
        if reference:
            key = (
                fee["tenant_id"],
                fee["account_id"],
                fee["currency"],
                reference,
            )
            fee_index.setdefault(key, []).append(fee)
    for bucket in fee_index.values():
        bucket.sort(key=lambda tx: (tx["booked_at_utc"], tx["txn_uid"]))
    result: dict[str, dict[str, Any]] = {}
    pair_checks = 0
    max_bucket_size = 0
    for outflow in principals:
        if (
            outflow["direction"] != "OUTFLOW"
            or outflow["counterparty_resolution"] != "EXACT_VERIFIED"
            or outflow["resolved_counterparty_account_id"] is None
        ):
            continue
        amount_rules: dict[int, dict[str, Any] | None] = {outflow["amount_minor"]: None}
        at = parse_time(outflow["booked_at_utc"])
        for rule in policy["fee_rules"]:
            if not isinstance(rule, dict) or type(rule.get("fee_minor")) is not int:
                block("BLOCK_POLICY")
            principal_amount = outflow["amount_minor"] - rule["fee_minor"]
            if principal_amount <= 0:
                continue
            if (
                rule.get("bank_code") == outflow["bank_code"]
                and rule.get("product_code") == outflow["product_code_id"]
                and rule.get("channel") == outflow["channel_id"]
                and type(rule.get("min_amount_minor")) is int
                and rule["min_amount_minor"]
                <= principal_amount
                <= rule.get("max_amount_minor", -1)
                and parse_time(rule["effective_from"])
                <= at
                <= parse_time(rule["effective_to"])
            ):
                if (
                    principal_amount in amount_rules
                    and amount_rules[principal_amount] is not None
                ):
                    block("BLOCK_AMBIGUOUS_FEE_RULE")
                amount_rules[principal_amount] = rule
        for principal_amount in sorted(amount_rules):
            bucket = inflow_index.get(
                (
                    outflow["tenant_id"],
                    outflow["resolved_counterparty_account_id"],
                    outflow["currency"],
                    principal_amount,
                ),
                [],
            )
            max_bucket_size = max(max_bucket_size, len(bucket))
            if len(bucket) > policy["max_trace_bucket"]:
                block("BLOCK_RESOURCE_TRACE_BUCKET")
            for inflow in bucket:
                pair_checks += 1
                if pair_checks > policy["max_pair_checks"]:
                    block("BLOCK_RESOURCE_PAIR_CHECKS")
                if (
                    inflow["counterparty_resolution"] != "EXACT_VERIFIED"
                    or inflow["resolved_counterparty_account_id"]
                    != outflow["account_id"]
                    or outflow["account_id"] == inflow["account_id"]
                ):
                    continue
                left_day = datetime.fromisoformat(
                    outflow.get("value_date") or outflow["local_date"]
                ).date()
                right_day = datetime.fromisoformat(
                    inflow.get("value_date") or inflow["local_date"]
                ).date()
                gap = (right_day - left_day).days
                if (
                    not policy["min_signed_gap_days"]
                    <= gap
                    <= policy["max_signed_gap_days"]
                ):
                    continue
                time_delta = int(
                    abs(
                        (
                            parse_time(inflow["booked_at_utc"])
                            - parse_time(outflow["booked_at_utc"])
                        ).total_seconds()
                    )
                    * 1000
                )
                basis = "RECIPROCAL_ACCOUNT"
                binding = digest(
                    {
                        "domain": "BUTLER_A4_STRONG_BINDING_V2_1",
                        "tenant_id": outflow["tenant_id"],
                        "outflow_account_id": outflow["account_id"],
                        "inflow_account_id": inflow["account_id"],
                        "outflow_resolved_counterparty_account_id": outflow[
                            "resolved_counterparty_account_id"
                        ],
                        "inflow_resolved_counterparty_account_id": inflow[
                            "resolved_counterparty_account_id"
                        ],
                        "eligibility_basis": basis,
                    }
                )
                candidates: list[dict[str, Any]] = []
                common = {
                    "outflow_txn_uid": outflow["txn_uid"],
                    "inflow_txn_uid": inflow["txn_uid"],
                    "currency": outflow["currency"],
                    "time_delta_ms": time_delta,
                    "exact_reference": exact_ref(outflow, inflow),
                    "eligibility_basis": basis,
                    "binding_digest": binding,
                    "affects_reporting": False,
                }
                if outflow["amount_minor"] == inflow["amount_minor"]:
                    fee_bucket = (
                        fee_index.get(
                            (
                                outflow["tenant_id"],
                                outflow["account_id"],
                                outflow["currency"],
                                outflow["bank_reference_token"],
                            ),
                            [],
                        )
                        if outflow.get("bank_reference_token")
                        else []
                    )
                    matching_fees = [
                        fee
                        for fee in fee_bucket
                        if fee["direction"] == "OUTFLOW"
                        and abs(
                            (
                                parse_time(fee["booked_at_utc"])
                                - parse_time(outflow["booked_at_utc"])
                            ).total_seconds()
                        )
                        <= 86400
                    ]
                    if matching_fees:
                        for fee in matching_fees:
                            candidates.append(
                                {
                                    **common,
                                    "fee_txn_uid": fee["txn_uid"],
                                    "fee_mode": "SEPARATE_FEE_ROW",
                                    "fee_minor": fee["amount_minor"],
                                    "evidence_grade": "OBSERVED_FEE",
                                    "policy_rule_id": None,
                                }
                            )
                    else:
                        candidates.append(
                            {
                                **common,
                                "fee_txn_uid": None,
                                "fee_mode": "NO_FEE",
                                "fee_minor": 0,
                                "evidence_grade": "OBSERVED_PRINCIPAL",
                                "policy_rule_id": None,
                            }
                        )
                elif outflow["amount_minor"] > inflow["amount_minor"]:
                    delta = outflow["amount_minor"] - inflow["amount_minor"]
                    rule = amount_rules.get(inflow["amount_minor"])
                    if rule is not None and rule["fee_minor"] == delta:
                        candidates.append(
                            {
                                **common,
                                "fee_txn_uid": None,
                                "fee_mode": "EMBEDDED_DEBIT",
                                "fee_minor": delta,
                                "evidence_grade": "SCHEDULE_SUPPORTED_FEE_CANDIDATE",
                                "policy_rule_id": rule["rule_id"],
                            }
                        )
                for candidate in candidates:
                    candidate["edge_id"] = edge_digest(candidate)
                    result[candidate["edge_id"]] = candidate
                    if len(result) > policy["max_candidate_edges"]:
                        block("BLOCK_RESOURCE_LIMIT")
    outflows = [tx for tx in principals if tx["direction"] == "OUTFLOW"]
    inflows = [tx for tx in principals if tx["direction"] == "INFLOW"]
    subset_branches = 0

    def add_subset(
        candidate_outflows: tuple[dict[str, Any], ...],
        candidate_inflows: tuple[dict[str, Any], ...],
    ) -> None:
        nonlocal subset_branches
        subset_branches += 1
        if subset_branches > branch_cap:
            block("BLOCK_RESOURCE_SUBSET_BRANCHES")
        currencies = {
            tx["currency"] for tx in (*candidate_outflows, *candidate_inflows)
        }
        if len(currencies) != 1:
            return
        out_total = sum(tx["amount_minor"] for tx in candidate_outflows)
        in_total = sum(tx["amount_minor"] for tx in candidate_inflows)
        fee_rule = None
        if out_total == in_total:
            match_type = "PARTIAL"
            fee_mode, fee_minor = "NO_FEE", 0
            grade = "OBSERVED_PRINCIPAL"
        elif out_total > in_total:
            fee_minor = out_total - in_total
            matching_rules = []
            for rule in policy["fee_rules"]:
                if not isinstance(rule, dict) or type(rule.get("fee_minor")) is not int:
                    block("BLOCK_POLICY")
                if rule["fee_minor"] != fee_minor:
                    continue
                if all(
                    rule.get("bank_code") == tx["bank_code"]
                    and rule.get("product_code") == tx["product_code_id"]
                    and rule.get("channel") == tx["channel_id"]
                    and type(rule.get("min_amount_minor")) is int
                    and rule["min_amount_minor"] <= in_total <= rule.get("max_amount_minor", -1)
                    and parse_time(rule["effective_from"])
                    <= parse_time(tx["booked_at_utc"])
                    <= parse_time(rule["effective_to"])
                    for tx in candidate_outflows
                ):
                    matching_rules.append(rule)
            if len(matching_rules) > 1:
                block("BLOCK_AMBIGUOUS_FEE_RULE")
            if not matching_rules:
                return
            fee_rule = matching_rules[0]
            match_type, fee_mode = "FEE_ADJUSTED", "EMBEDDED_DEBIT"
            grade = "SCHEDULE_SUPPORTED_FEE_CANDIDATE"
        else:
            return
        ordered_out = tuple(sorted(candidate_outflows, key=lambda tx: tx["txn_uid"]))
        ordered_in = tuple(sorted(candidate_inflows, key=lambda tx: tx["txn_uid"]))
        edge = {
            "outflow_txn_uid": ordered_out[0]["txn_uid"],
            "inflow_txn_uid": ordered_in[0]["txn_uid"],
            "fee_txn_uid": None,
            "fee_mode": fee_mode,
            "fee_minor": fee_minor,
            "currency": next(iter(currencies)),
            "evidence_grade": grade,
            "policy_rule_id": fee_rule.get("rule_id") if fee_rule else None,
            "time_delta_ms": aggregate_delta_ms((*ordered_out, *ordered_in)),
            "exact_reference": all(
                exact_ref(left, right) for left in ordered_out for right in ordered_in
            ),
            "eligibility_basis": "RECIPROCAL_ACCOUNT",
            "binding_digest": multi_binding(ordered_out, ordered_in, match_type),
            "affects_reporting": False,
            "outflow_txn_uids": [tx["txn_uid"] for tx in ordered_out],
            "inflow_txn_uids": [tx["txn_uid"] for tx in ordered_in],
            "match_type": match_type,
            "currency_mode": next(iter(currencies)),
            "fx_rule_id": None,
            "fx_receipt_digest": None,
        }
        edge["edge_id"] = edge_digest(edge)
        result[edge["edge_id"]] = edge
        if len(result) > policy["max_candidate_edges"]:
            block("BLOCK_RESOURCE_LIMIT")

    for outflow in outflows:
        eligible = sorted(
            (
                tx for tx in inflows
                if tx["currency"] == outflow["currency"]
                and strong_pair(outflow, tx, policy)
            ),
            key=lambda tx: tx["txn_uid"],
        )
        for size in range(2, min(subset_cap, len(eligible)) + 1):
            for subset in itertools.combinations(eligible, size):
                add_subset((outflow,), subset)
    for inflow in inflows:
        eligible = sorted(
            (
                tx for tx in outflows
                if tx["currency"] == inflow["currency"]
                and strong_pair(tx, inflow, policy)
            ),
            key=lambda tx: tx["txn_uid"],
        )
        for size in range(2, min(subset_cap, len(eligible)) + 1):
            for subset in itertools.combinations(eligible, size):
                add_subset(subset, (inflow,))

    for outflow in outflows:
        for inflow in inflows:
            if outflow["currency"] == inflow["currency"] or not strong_pair(
                outflow, inflow, policy
            ):
                continue
            matching = [
                rule for rule in policy.get("fx_rules", [])
                if exact_fx_rule(rule, outflow, inflow)
            ]
            if len(matching) > 1:
                block("BLOCK_AMBIGUOUS_FX_RULE")
            if not matching:
                continue
            rule = matching[0]
            edge = {
                "outflow_txn_uid": outflow["txn_uid"],
                "inflow_txn_uid": inflow["txn_uid"],
                "fee_txn_uid": None,
                "fee_mode": "NO_FEE",
                "fee_minor": 0,
                "currency": outflow["currency"],
                "evidence_grade": "RECEIPT_BOUND_FX",
                "policy_rule_id": rule["rule_id"],
                "time_delta_ms": aggregate_delta_ms((outflow, inflow)),
                "exact_reference": exact_ref(outflow, inflow),
                "eligibility_basis": "RECIPROCAL_ACCOUNT",
                "binding_digest": multi_binding((outflow,), (inflow,), "FX", rule),
                "affects_reporting": False,
                "outflow_txn_uids": [outflow["txn_uid"]],
                "inflow_txn_uids": [inflow["txn_uid"]],
                "match_type": "FX",
                "currency_mode": f"{rule['from_currency']}/{rule['to_currency']}",
                "fx_rule_id": rule["rule_id"],
                "fx_receipt_digest": rule["rate_source_receipt_digest"],
            }
            edge["edge_id"] = edge_digest(edge)
            result[edge["edge_id"]] = edge
            if len(result) > policy["max_candidate_edges"]:
                block("BLOCK_RESOURCE_LIMIT")
    return [result[key] for key in sorted(result)], pair_checks, max_bucket_size, subset_branches


def edge_nodes(edge: Mapping[str, Any]) -> frozenset[str]:
    values = {
        edge["outflow_txn_uid"],
        edge["inflow_txn_uid"],
        edge["fee_txn_uid"],
        *edge.get("outflow_txn_uids", []),
        *edge.get("inflow_txn_uids", []),
    }
    return frozenset(value for value in values if value)


def components(edges: Sequence[dict[str, Any]]) -> list[list[dict[str, Any]]]:
    by_node: dict[str, list[dict[str, Any]]] = {}
    for edge in edges:
        for node in edge_nodes(edge):
            by_node.setdefault(node, []).append(edge)
    unseen = set(by_node)
    output = []
    while unseen:
        stack, found = [min(unseen)], {}
        unseen.remove(stack[0])
        while stack:
            node = stack.pop()
            for edge in by_node[node]:
                found[edge["edge_id"]] = edge
                for other in edge_nodes(edge):
                    if other in unseen:
                        unseen.remove(other)
                        stack.append(other)
        output.append([found[key] for key in sorted(found)])
    return sorted(output, key=lambda group: group[0]["edge_id"])


@dataclass(frozen=True)
class DP:
    objective: tuple[int, int, int]
    count: int
    union: frozenset[str]
    intersection: frozenset[str]
    representative: tuple[str, ...]


def solve(group: Sequence[dict[str, Any]]) -> DP:
    outflows = sorted({edge["outflow_txn_uid"] for edge in group})
    by_out = {
        uid: [edge for edge in group if edge["outflow_txn_uid"] == uid]
        for uid in outflows
    }

    @lru_cache(maxsize=None)
    def visit(index: int, used: frozenset[str]) -> DP:
        if index == len(outflows):
            return DP((0, 0, 0), 1, frozenset(), frozenset(), ())
        values = [visit(index + 1, used)]
        for edge in by_out[outflows[index]]:
            nodes = edge_nodes(edge)
            if nodes & used:
                continue
            suffix = visit(index + 1, used | nodes)
            cost = (
                -len(edge_nodes(edge) - ({edge["fee_txn_uid"]} if edge["fee_txn_uid"] else set())),
                edge["time_delta_ms"],
                int(edge["evidence_grade"] == "SCHEDULE_SUPPORTED_FEE_CANDIDATE"),
            )
            edge_set = frozenset({edge["edge_id"]})
            values.append(
                DP(
                    tuple(a + b for a, b in zip(cost, suffix.objective)),
                    suffix.count,
                    suffix.union | edge_set,
                    suffix.intersection | edge_set,
                    tuple(sorted((*suffix.representative, edge["edge_id"]))),
                )
            )
        best = min(value.objective for value in values)
        winners = [value for value in values if value.objective == best]
        union = frozenset().union(*(value.union for value in winners))
        intersection = winners[0].intersection
        for winner in winners[1:]:
            intersection &= winner.intersection
        return DP(
            best,
            sum(value.count for value in winners),
            union,
            intersection,
            min(value.representative for value in winners),
        )

    return visit(0, frozenset())


def expected_classification(
    run_id: str,
    txs: Sequence[dict[str, Any]],
    edges: list[dict[str, Any]],
    terminal_groups: Sequence[dict[str, Any]] = (),
) -> dict[str, Any]:
    component_docs: list[dict[str, Any]] = []
    forced_all: set[str] = set()
    graph_nodes: set[str] = set()
    for group in components(edges):
        nodes = frozenset().union(*(edge_nodes(edge) for edge in group))
        graph_nodes |= nodes
        value = solve(group)
        forced_all |= value.intersection
        nonforced = value.union - value.intersection
        ambiguous = sorted(
            frozenset().union(
                *(edge_nodes(edge) for edge in group if edge["edge_id"] in nonforced)
            )
            if nonforced
            else frozenset()
        )
        selectable = (
            frozenset().union(
                *(edge_nodes(edge) for edge in group if edge["edge_id"] in value.union)
            )
            if value.union
            else frozenset()
        )
        component_id = digest(group)
        exact = len(nodes) <= 8
        component_docs.append(
            {
                "run_id": run_id,
                "component_id": component_id,
                "forced_edge_ids": sorted(value.intersection),
                "ambiguous_txn_uids": ambiguous,
                "unmatched_txn_uids": sorted(nodes - selectable),
                "has_alternative_optimum": value.count > 1,
                "alternative_optimum_count": value.count if exact else None,
                "alternative_optimum_count_exact": exact,
                "forced_pair_count": len(value.intersection),
                "residual_unmatched_count": len(nodes - selectable),
            }
        )
    forced_edges = [edge for edge in edges if edge["edge_id"] in forced_all]
    unmatched = sorted(tx["txn_uid"] for tx in txs if tx["txn_uid"] not in graph_nodes)
    result = {
        "schema_version": "2.0.0",
        "run_id": run_id,
        "forced_edges": forced_edges,
        "components": sorted(component_docs, key=lambda item: item["component_id"]),
        "globally_unmatched_txn_uids": unmatched,
        "shadow_only": True,
        "affects_reporting": False,
    }
    if terminal_groups:
        result["terminal_groups"] = list(terminal_groups)
    return result


def verify_dlp(docs: Mapping[str, Any]) -> None:
    forbidden_key = re.compile(
        r"(?:^|_)(?:raw_text|raw_value|memo|holder_name|person_name|file_path|secret|password|account_number)(?:$|_)",
        re.I,
    )
    account = re.compile(r"(?<![A-Za-z0-9])(?:[0-9０-９][ -]?){8,20}(?![A-Za-z0-9])")
    home = re.compile(r"(?:/Users/|/home/|[A-Za-z]:\\\\Users\\\\)")
    secret = re.compile(
        r"(?:bearer\s+[A-Za-z0-9._~-]+|api[_-]?key|private[_-]?key)", re.I
    )
    email = re.compile(r"(?i)(?<![A-Z0-9._%+-])[A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,}(?![A-Z0-9.-])")
    allowed = re.compile(r"^[A-Za-z0-9._:+/@=-]{0,256}$")
    machine_identifier = re.compile(
        r"^(?:[0-9a-f]{32}|[0-9a-f]{40}|[0-9a-f]{64}|"
        r"[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12}|"
        r"[0-9]{4}-[0-9]{2}-[0-9]{2}(?:T[0-9:.+-]+Z?)?)$"
    )
    adapter_header_keys = {
        "source_account",
        "bank_code",
        "booking_date",
        "withdrawal",
        "deposit",
        "reference",
        "counter_account",
        "fee_type",
        "product_code",
        "channel",
        "currency",
        "value_date",
        "duplicate_of_reference",
        "reversal_of_reference",
        "allowed_date_formats",
    }
    count = 0
    protected_machine_keys = re.compile(
        r"(?:digest|token|signature|sha256|nonce|oid|uuid|_id)$", re.I
    )

    def scan_text(text: str, key: str, depth: int = 0) -> None:
        if protected_machine_keys.search(key):
            return
        if machine_identifier.fullmatch(text) and not re.fullmatch(
            r"[0-9a-f]{32,64}", text
        ):
            return
        if home.search(text) or secret.search(text) or email.search(text) or account.search(text):
            block("BLOCK_DLP")
        if depth >= 2:
            return
        decoded: list[bytes] = []
        if re.fullmatch(r"[0-9A-Fa-f]{8,512}", text) and len(text) % 2 == 0:
            try:
                decoded.append(bytes.fromhex(text))
            except ValueError:
                pass
        if re.fullmatch(r"[A-Za-z0-9_-]{8,684}={0,2}", text):
            try:
                padded = text + "=" * (-len(text) % 4)
                decoded.append(base64.b64decode(padded, altchars=b"-_", validate=True))
            except (ValueError, TypeError):
                pass
        for raw_value in decoded:
            try:
                nested = raw_value.decode("utf-8")
            except UnicodeDecodeError:
                continue
            if nested != text:
                scan_text(nested, key, depth + 1)

    def visit(value: Any, key: str = "") -> None:
        nonlocal count
        if forbidden_key.search(key):
            block("BLOCK_DLP")
        if isinstance(value, dict):
            for child_key, child in value.items():
                visit(child, child_key)
        elif isinstance(value, list):
            for child in value:
                visit(child, key)
        elif isinstance(value, str):
            count += 1
            scan_text(value, key)
            if machine_identifier.fullmatch(value):
                return
            if key in adapter_header_keys and len(value.encode("utf-8")) <= 96:
                return
            if not allowed.fullmatch(value):
                block("BLOCK_DLP")
        elif value is not None and type(value) not in {int, bool}:
            block("BLOCK_DLP")

    scanned = {
        name: docs[name] for name in JSON_FILES - {"dlp_receipt.json", "manifest.json"}
    }
    for document in scanned.values():
        visit(document)
    receipt = docs["dlp_receipt.json"]
    if (
        receipt.get("files_scanned") != len(scanned)
        or receipt.get("values_scanned") != count
        or receipt.get("leak_count") != 0
        or receipt.get("closed_world_pass") is not True
    ):
        block("BLOCK_DLP")


def verify(
    root: Path,
    trust_path: Path,
    source_fd: int,
    source_suffix: str,
    key_material: Mapping[str, Any],
    dictionary_path: Path,
    *,
    require_bundled_build_stamp: bool = False,
) -> dict[str, Any]:
    raw = read_files(root)
    docs: dict[str, Any] = {}
    for name in JSON_FILES:
        docs[name] = strict_load(raw[name])
        if jcs(docs[name]) != raw[name]:
            block("BLOCK_CANONICAL_JSON")
    verify_hashes(raw, docs)
    run, txs, expected = verify_bindings(docs)
    verify_source_snapshot(source_fd, source_suffix, docs["input_receipt.json"])
    material = _key_material(key_material, run["tenant_id"])
    dictionary = _dictionary(dictionary_path, run["dictionary_digest"])
    frame = read_source_frame(source_fd, source_suffix)
    independently_compiled = independently_compile_source(
        frame,
        docs["input_receipt.json"],
        run,
        docs["account_snapshot_receipt.json"]["accounts"],
        material,
        dictionary,
    )
    if independently_compiled != txs:
        block("BLOCK_SOURCE_SEMANTIC_MISMATCH")
    trust = load_trust(trust_path)
    graph = docs["graph_observation.json"]
    policy = verify_policy(
        docs["policy_receipt.json"], graph.get("policy_payload"), trust, expected
    )
    if docs["policy_receipt.json"]["policy_digest"] != run["policy_digest"]:
        block("BLOCK_BINDING")
    active_txs, terminal_groups = terminal_partition(txs)
    rebuilt, pair_checks, max_bucket_size, subset_branches = rebuild_edges(
        active_txs, policy
    )
    observed = graph
    expected_graph_keys = {
        "schema_version",
        "run_id",
        "nonce_digest",
        "tenant_id",
        "objective",
        "policy_payload",
        "candidate_pair_checks",
        "max_candidate_bucket_size",
        "candidate_subset_branches",
        "edges",
    }
    if (
        set(observed) != expected_graph_keys
        or observed.get("schema_version") != "2.1.0"
        or observed.get("objective")
        != [
            "MAX_MATCHED_PRINCIPALS",
            "MIN_TIME_DISTANCE",
            "MIN_FEE_PENALTY",
            "ORDERED_EDGE_ID_VECTOR",
        ]
        or observed.get("candidate_pair_checks") != pair_checks
        or observed.get("max_candidate_bucket_size") != max_bucket_size
        or observed.get("candidate_subset_branches") != subset_branches
        or observed.get("edges") != rebuilt
    ):
        block("BLOCK_GRAPH")
    expected_result = expected_classification(
        run["run_id"], active_txs, rebuilt, terminal_groups
    )
    if docs["classification_receipt.json"] != expected_result:
        block("BLOCK_OPTIMUM")
    invariants = docs["runtime_invariants.json"]
    exact_invariants = {
        "run_id": run["run_id"],
        "nonce_digest": expected["run_nonce_digest"],
        "shadow_only": True,
        "product_mutation_count": 0,
        "journal_auto_post_allowed": False,
        "affects_reporting": False,
        "runtime_activation_allowed": False,
        "observer_product_bytes_equal": True,
        "terminal_count": 1,
    }
    if invariants != exact_invariants:
        block("BLOCK_INVARIANT")
    verify_dlp(docs)
    resources_root = Path(__file__).resolve().parents[2]
    closure_files: dict[str, str] = {}
    for relative in CODE_FILES:
        target = resources_root / relative
        if not target.is_file() or target.is_symlink():
            block("BLOCK_CODE_CLOSURE")
        closure_files[relative] = hashlib.sha256(target.read_bytes()).hexdigest()
    code_closure_digest = digest(closure_files)
    if require_bundled_build_stamp:
        stamp_path = resources_root / "BUILD_INFO.json"
        try:
            stamp = strict_load(_read_trust_file(stamp_path))
        except Block:
            block("BLOCK_CODE_CLOSURE")
        stamped_closure = stamp.get("a4_code_closure") if isinstance(stamp, dict) else None
        if (
            not isinstance(stamped_closure, dict)
            or stamp.get("build_tree_oid") != run["code_tree_oid"]
            or stamped_closure.get("files") != closure_files
            or stamped_closure.get("digest") != code_closure_digest
        ):
            block("BLOCK_CODE_CLOSURE")
    verified_at = datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")
    receipt = {
        "schema_version": "butler.box5.a4.verification_receipt.v5.3",
        "contract_id": "BUTLER-BOX5-A4-V5.3-NATIVE-AUTHORITY",
        "run_id": run["run_id"],
        "nonce": run["nonce"].lower(),
        "decision": "PASS",
        "reason_codes": [],
        "source_sha256": docs["input_receipt.json"]["source_file_sha256"],
        "ordered_row_root": docs["input_receipt.json"]["row_closure"]["ordered_row_root"],
        "compiled_manifest_digest": digest(independently_compiled),
        "graph_digest": digest(graph),
        "evidence_manifest_digest": hashlib.sha256(raw["manifest.json"]).hexdigest(),
        "policy_digest": run["policy_digest"],
        "adapter_digest": run["adapter_digest"],
        "dictionary_digest": run["dictionary_digest"],
        "registry_digest": run["registry_digest"],
        "code_tree_oid": run["code_tree_oid"],
        "code_closure_digest": code_closure_digest,
        "verifier_id": "BUTLER_A4_ISOLATED_RAW_COMPILER_V5.3",
        "verified_at_utc": verified_at,
    }
    return receipt


def _decode_b64(value: object, *, maximum: int) -> bytes:
    if not isinstance(value, str) or len(value) > (maximum * 4 // 3) + 8:
        block("BLOCK_RESOURCE_LIMIT")
    try:
        decoded = base64.b64decode(value, validate=True)
    except (ValueError, TypeError):
        block("BLOCK_SCHEMA")
    if not 0 < len(decoded) <= maximum:
        block("BLOCK_RESOURCE_LIMIT")
    return decoded


def _read_trust_file(path: Path) -> bytes:
    flags = os.O_RDONLY | getattr(os, "O_CLOEXEC", 0) | getattr(os, "O_NOFOLLOW", 0)
    descriptor = -1
    try:
        descriptor = os.open(path, flags)
        before = os.fstat(descriptor)
        if (
            not stat.S_ISREG(before.st_mode)
            or before.st_mode & 0o022
            or not 0 < before.st_size <= 16_384
        ):
            block("BLOCK_TRUST_UNAVAILABLE")
        chunks: list[bytes] = []
        remaining = 16_385
        while remaining:
            chunk = os.read(descriptor, min(remaining, 4096))
            if not chunk:
                break
            chunks.append(chunk)
            remaining -= len(chunk)
        raw = b"".join(chunks)
        after = os.fstat(descriptor)
        before_identity = (
            before.st_dev,
            before.st_ino,
            before.st_size,
            before.st_mtime_ns,
            before.st_ctime_ns,
        )
        after_identity = (
            after.st_dev,
            after.st_ino,
            after.st_size,
            after.st_mtime_ns,
            after.st_ctime_ns,
        )
        if not 0 < len(raw) <= 16_384 or before_identity != after_identity:
            block("BLOCK_TRUST_UNAVAILABLE")
        return raw
    except Block:
        raise
    except OSError:
        block("BLOCK_TRUST_UNAVAILABLE")
    finally:
        if descriptor >= 0:
            os.close(descriptor)


def _write_private(path: Path, payload: bytes) -> None:
    descriptor = os.open(
        path,
        os.O_WRONLY | os.O_CREAT | os.O_EXCL | getattr(os, "O_CLOEXEC", 0),
        0o600,
    )
    try:
        view = memoryview(payload)
        while view:
            written = os.write(descriptor, view)
            if written <= 0:
                block("BLOCK_INPUT_CLOSURE")
            view = view[written:]
        os.fsync(descriptor)
    finally:
        os.close(descriptor)


def _verified_unsigned_receipt(
    request: Mapping[str, Any],
    source_fd: int,
    *,
    require_bundled_build_stamp: bool,
) -> dict[str, Any]:
    required = {
        "schema_version",
        "protocol_version",
        "tenant_id",
        "source_suffix",
        "authority_trust_digest",
        "source_payload_digest",
        "evidence_files_b64",
        "finance_trust_b64",
        "dictionary_b64",
        "key_material",
    }
    if (
        set(request) != required
        or request.get("schema_version") != "butler.box5.a4.producer_request.v5.3"
        or request.get("protocol_version") != PROTOCOL_VERSION
        or not re.fullmatch(
            r"[0-9a-f]{64}", str(request.get("authority_trust_digest", ""))
        )
        or not re.fullmatch(
            r"[0-9a-f]{64}", str(request.get("source_payload_digest", ""))
        )
        or request.get("source_suffix") not in {".csv", ".xls", ".xlsx"}
        or not isinstance(request.get("evidence_files_b64"), dict)
        or set(request["evidence_files_b64"]) != FILES
        or not isinstance(request.get("key_material"), dict)
    ):
        block("BLOCK_SCHEMA")
    evidence = {
        name: _decode_b64(value, maximum=8 * 1024 * 1024)
        for name, value in request["evidence_files_b64"].items()
    }
    run = strict_load(evidence["run_request.json"])
    if (
        not isinstance(run, dict)
        or run.get("tenant_id") != request.get("tenant_id")
        or not re.fullmatch(
            r"[0-9a-f]{8}-[0-9a-f]{4}-[1-5][0-9a-f]{3}-[89ab][0-9a-f]{3}-[0-9a-f]{12}",
            str(run.get("run_id", "")),
            re.I,
        )
    ):
        block("BLOCK_RUN_BINDING")
    finance_trust = _decode_b64(request.get("finance_trust_b64"), maximum=64 * 1024)
    dictionary = _decode_b64(request.get("dictionary_b64"), maximum=4 * 1024 * 1024)
    with tempfile.TemporaryDirectory(prefix="butler-a4-authority-") as temporary:
        temporary_root = Path(temporary)
        evidence_root = temporary_root / str(run["run_id"])
        evidence_root.mkdir(mode=0o700)
        for name, payload in evidence.items():
            _write_private(evidence_root / name, payload)
        finance_path = temporary_root / "finance-trust.json"
        dictionary_path = temporary_root / "dictionary.json"
        _write_private(finance_path, finance_trust)
        _write_private(dictionary_path, dictionary)
        return verify(
            evidence_root,
            finance_path,
            source_fd,
            str(request["source_suffix"]),
            request["key_material"],
            dictionary_path,
            require_bundled_build_stamp=require_bundled_build_stamp,
        )


def _read_stream_exact(size: int) -> bytes:
    if size < 0:
        block("BLOCK_REQUEST_SIZE")
    value = sys.stdin.buffer.read(size)
    if len(value) != size:
        block("BLOCK_REQUEST_SIZE")
    return value


def _receive_verify_request() -> tuple[dict[str, Any], bytes]:
    request_size = struct.unpack(">I", _read_stream_exact(4))[0]
    if not 0 < request_size <= MAX_REQUEST_BYTES:
        block("BLOCK_REQUEST_SIZE")
    wrapper = strict_load(_read_stream_exact(request_size))
    source_size = struct.unpack(">Q", _read_stream_exact(8))[0]
    if not 0 < source_size <= MAX_SOURCE_BYTES:
        block("BLOCK_REQUEST_SIZE")
    source = _read_stream_exact(source_size)
    if sys.stdin.buffer.read(1):
        block("BLOCK_REQUEST_SCHEMA")
    if not isinstance(wrapper, dict):
        block("BLOCK_REQUEST_SCHEMA")
    return wrapper, source


def _verify_request_wrapper(
    wrapper: Mapping[str, Any], source: bytes, *, require_bundled_build_stamp: bool
) -> dict[str, Any]:
    required = {
        "schema_version",
        "protocol_version",
        "challenge",
        "producer_request",
        "request_digest",
        "source_payload_digest",
    }
    if (
        set(wrapper) != required
        or wrapper.get("schema_version") != VERIFY_REQUEST_SCHEMA
        or wrapper.get("protocol_version") != PROTOCOL_VERSION
        or not isinstance(wrapper.get("challenge"), dict)
        or not isinstance(wrapper.get("producer_request"), dict)
    ):
        block("BLOCK_REQUEST_SCHEMA")
    try:
        challenge = AuthorityChallenge.from_mapping(wrapper["challenge"])
        computed_request_digest = request_digest(wrapper["producer_request"])
    except (TypeError, ValueError):
        block("BLOCK_REQUEST_SCHEMA")
    source_digest = sha256_hex(source)
    if (
        wrapper.get("request_digest") != computed_request_digest
        or wrapper.get("source_payload_digest") != source_digest
        or wrapper["producer_request"].get("source_payload_digest") != source_digest
    ):
        block("BLOCK_SOURCE_DIGEST")
    source_descriptor = -1
    with tempfile.TemporaryDirectory(prefix="butler-a4-verifier-") as temporary:
        source_path = Path(temporary) / ("source" + str(wrapper["producer_request"].get("source_suffix", "")))
        _write_private(source_path, source)
        source_descriptor = os.open(
            source_path,
            os.O_RDONLY | getattr(os, "O_CLOEXEC", 0) | getattr(os, "O_NOFOLLOW", 0),
        )
        try:
            receipt = _verified_unsigned_receipt(
                wrapper["producer_request"],
                source_descriptor,
                require_bundled_build_stamp=require_bundled_build_stamp,
            )
        finally:
            os.close(source_descriptor)
    return verified_observation(
        receipt,
        challenge=challenge,
        request_digest=computed_request_digest,
        source_payload_digest=source_digest,
    )


def _write_observation(observation: Mapping[str, Any]) -> None:
    payload = jcs(observation)
    if not 0 < len(payload) <= MAX_RESPONSE_BYTES:
        block("BLOCK_REQUEST_SIZE")
    sys.stdout.buffer.write(struct.pack(">I", len(payload)))
    sys.stdout.buffer.write(payload)
    sys.stdout.buffer.flush()


def main() -> int:
    parser = argparse.ArgumentParser(add_help=False)
    parser.add_argument("--verify-observation", action="store_true")
    parser.add_argument("--require-bundled-build-stamp", action="store_true")
    try:
        args = parser.parse_args()
        if not args.verify_observation:
            block("BLOCK_PROTOCOL_VERSION")
        wrapper, source = _receive_verify_request()
        _write_observation(
            _verify_request_wrapper(
                wrapper,
                source,
                require_bundled_build_stamp=args.require_bundled_build_stamp,
            )
        )
        return 0
    except Block as exc:
        _write_observation(
            {
                "schema_version": OBSERVATION_SCHEMA,
                "protocol_version": PROTOCOL_VERSION,
                "status": "BLOCK",
                "error_code": public_error_code(str(exc)),
                "authority_session_id": None,
                "authority_request_id": None,
                "authority_nonce_digest": None,
                "request_digest": None,
                "source_payload_digest": None,
                "receipt": None,
            }
        )
        return 1
    except Exception:
        _write_observation(
            {
                "schema_version": OBSERVATION_SCHEMA,
                "protocol_version": PROTOCOL_VERSION,
                "status": "BLOCK",
                "error_code": "UNVERIFIED_INTERNAL",
                "authority_session_id": None,
                "authority_request_id": None,
                "authority_nonce_digest": None,
                "request_digest": None,
                "source_payload_digest": None,
                "receipt": None,
            }
        )
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
