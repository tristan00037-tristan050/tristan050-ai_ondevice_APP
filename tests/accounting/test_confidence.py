"""test_confidence.py — 신뢰도 범위(50~100%) 및 정수% 표시 형식 검증."""
from __future__ import annotations

import re
import tempfile
from pathlib import Path

import pytest

try:
    import pandas as pd
    import openpyxl
    _DEPS_OK = True
except ImportError:
    _DEPS_OK = False

_skip = pytest.mark.skipif(not _DEPS_OK, reason="pandas/openpyxl 미설치")


@_skip
def test_confidence_range_for_classified():
    """분류 성공 항목의 신뢰도가 0.50~1.00 범위여야 한다."""
    from butler_pc_core.accounting.classifier import classify_df
    from .conftest import make_synthetic_df

    df = classify_df(make_synthetic_df(50))
    classified = df[df["분류과목"] != "미분류"]

    assert not classified.empty, "분류된 항목 없음 — 테스트 데이터 확인 필요"
    low = classified[classified["신뢰도"] < 0.50]
    assert low.empty, (
        f"신뢰도 50% 미만인 분류 항목 발견:\n"
        + low[["적요", "분류과목", "신뢰도"]].to_string()
    )
    assert (classified["신뢰도"] <= 1.00).all(), "신뢰도 100% 초과 항목 발견"


@_skip
def test_confidence_format():
    """xlsx [분류결과] 시트의 신뢰도 컬럼이 '정수%' 형식이어야 한다 (예: '58%')."""
    from butler_pc_core.accounting.classifier import classify_df, save_classified

    df = classify_df(pd.DataFrame({
        "적요": ["급여 지급", "통신비 납부"],
        "거래처": ["", ""],
        "금액": [500000, 80000],
    }))

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        tmp = f.name
    try:
        save_classified(df, tmp)
        wb = openpyxl.load_workbook(tmp)
        ws = wb["분류결과"]
        headers = [cell.value for cell in ws[1]]
        conf_col = headers.index("신뢰도") + 1  # 1-indexed
        pattern = re.compile(r"^\d+%$")
        for row in ws.iter_rows(min_row=2, min_col=conf_col, max_col=conf_col):
            val = str(row[0].value)
            assert pattern.match(val), (
                f"신뢰도 표시 형식 오류 — 기대: '정수%', 실제: '{val}'"
            )
    finally:
        Path(tmp).unlink(missing_ok=True)
