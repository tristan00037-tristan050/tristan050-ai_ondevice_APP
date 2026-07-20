from __future__ import annotations

import base64
import hashlib
import os
import shutil
import socket
import subprocess
import sys
import tempfile
import time
import uuid
from datetime import datetime, timedelta, timezone
from pathlib import Path

import pandas as pd
import pytest
from fastapi import FastAPI
from fastapi.testclient import TestClient
from cryptography.hazmat.primitives.asymmetric.ed25519 import Ed25519PrivateKey
from jsonschema import Draft202012Validator

import butler_pc_core.accounting.assignment.router as assignment_routes
import butler_pc_core.accounting.assignment.store as assignment_store_module
from butler_pc_core.a4_verifier.cli import AUTHORITY_MAX_CLIENTS
from butler_pc_core.accounting.assignment.runtime import (
    AccountingReviewRuntime,
    set_accounting_review_runtime_for_tests,
)
from butler_pc_core.accounting.assignment.security import MemoryKeyStore, TokenService
from butler_pc_core.accounting.assignment.store import SQLiteAssignmentStore
from butler_pc_core.accounting.assignment.domain import AssignmentError
from butler_pc_core.accounting.classify.reconciliation_service_v2 import (
    SCHEMA_DIGEST,
    approved_split_adapter,
    queue_canonical_product_reconciliation,
    run_canonical_product_reconciliation,
)
from butler_pc_core.accounting.classify import reconciliation_service_v2 as service_v2
from butler_pc_core.accounting.classify.verifier_authority import (
    _receive_response,
    _send_request,
    load_authority_trust,
    request_authority_verification,
)
from butler_pc_core.accounting.classify.reconciliation_v2 import (
    A4ContractError,
    CompileReceipt,
    CompiledTransaction,
    CounterpartyResolution,
    Direction,
    EligibilityBasis,
    EvidenceGrade,
    FeeMode,
    FeeRule,
    TrustPolicy,
    VerifiedPolicy,
    closed_world_dlp,
    compile_dataframe,
    digest_object,
    jcs_bytes,
    reconcile,
    sha256_bytes,
    strict_json_loads,
    tenant_uuid,
    verify_signed_policy,
)
from butler_pc_core.accounting.classify.source_snapshot_v2_1 import (
    secure_source_snapshot,
)
from butler_pc_core.accounting.classify.shadow.runner import (
    run_and_write_product_shadow,
)
from butler_pc_core.company_profile.contracts import make_runtime_profile
from butler_pc_core.auth.capability_token import CapabilityTokenManager


pytestmark = pytest.mark.no_sidecar_token


NOW = datetime(2026, 7, 19, 6, 0, tzinfo=timezone.utc)
RUN_ID = "0c567f70-4e31-4d1e-b880-c737f57b2a38"
TREE = "a" * 40
SOURCE_SHA = hashlib.sha256(b"physical-bank-file-v2").hexdigest()
SOURCE_SIZE = 1_024


@pytest.fixture(scope="module", autouse=True)
def isolated_verifier_authority(tmp_path_factory):
    """Run the real verifier authority with a key unavailable to product code."""

    # AF_UNIX paths are capped at roughly 104 bytes on macOS.  Keep the real
    # authority transport under a private, short-lived directory.
    del tmp_path_factory
    root = Path(tempfile.mkdtemp(prefix="butler-a4-authority-", dir="/tmp"))
    root.chmod(0o700)
    socket_path = root / "authority.sock"
    trust_path = root / "authority-trust.json"
    private_key = Ed25519PrivateKey.generate()
    trust = {
        "schema_version": "butler.box5.a4.verifier_authority_trust.v1",
        "enabled": True,
        "environment": "TEST",
        "authority_id": "butler-a4-isolated-test-authority",
        "transport": "unix-scm-rights-v1",
        "algorithm": "Ed25519",
        "signer_key_id": "isolated-test-authority-key-v1",
        "public_key_b64": base64.b64encode(
            private_key.public_key().public_bytes_raw()
        ).decode("ascii"),
        "revocation_epoch": 9,
        "not_before_utc": "2026-01-01T00:00:00Z",
        "not_after_utc": "2030-01-01T00:00:00Z",
    }
    trust_path.write_bytes(jcs_bytes(trust))
    trust_path.chmod(0o600)
    read_fd, write_fd = os.pipe()
    try:
        os.write(write_fd, private_key.private_bytes_raw())
    finally:
        os.close(write_fd)
    verifier = (
        Path(__file__).resolve().parents[3]
        / "butler_pc_core"
        / "a4_verifier"
        / "cli.py"
    )
    process = subprocess.Popen(  # nosec B603
        [
            sys.executable,
            "-I",
            str(verifier),
            "--serve-authority",
            "--socket",
            str(socket_path),
            "--authority-trust",
            str(trust_path),
            "--signing-key-fd",
            str(read_fd),
            "--allow-test-authority",
        ],
        cwd=str(Path(__file__).resolve().parents[3]),
        env={"PATH": os.environ.get("PATH", "")},
        pass_fds=(read_fd,),
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        text=False,
    )
    os.close(read_fd)
    deadline = time.monotonic() + 10
    while not socket_path.exists() and process.poll() is None:
        if time.monotonic() >= deadline:
            process.terminate()
            raise RuntimeError("isolated verifier authority did not start")
        time.sleep(0.02)
    if process.poll() is not None:
        stdout, stderr = process.communicate(timeout=1)
        raise RuntimeError(
            "isolated verifier authority failed: "
            f"stdout={stdout!r}, stderr={stderr!r}"
        )

    previous = (
        service_v2.VERIFIER_AUTHORITY_SOCKET_PATH,
        service_v2.VERIFIER_AUTHORITY_TRUST_PATH,
        service_v2.ALLOW_TEST_VERIFIER_AUTHORITY,
        assignment_store_module.A4_VERIFIER_AUTHORITY_TRUST_PATH,
        assignment_store_module.A4_ALLOW_TEST_VERIFIER_AUTHORITY,
    )
    service_v2.VERIFIER_AUTHORITY_SOCKET_PATH = socket_path
    service_v2.VERIFIER_AUTHORITY_TRUST_PATH = trust_path
    service_v2.ALLOW_TEST_VERIFIER_AUTHORITY = True
    assignment_store_module.A4_VERIFIER_AUTHORITY_TRUST_PATH = trust_path
    assignment_store_module.A4_ALLOW_TEST_VERIFIER_AUTHORITY = True
    try:
        yield {"trust_path": trust_path, "socket_path": socket_path, "trust": trust}
    finally:
        (
            service_v2.VERIFIER_AUTHORITY_SOCKET_PATH,
            service_v2.VERIFIER_AUTHORITY_TRUST_PATH,
            service_v2.ALLOW_TEST_VERIFIER_AUTHORITY,
            assignment_store_module.A4_VERIFIER_AUTHORITY_TRUST_PATH,
            assignment_store_module.A4_ALLOW_TEST_VERIFIER_AUTHORITY,
        ) = previous
        process.terminate()
        try:
            process.wait(timeout=5)
        except subprocess.TimeoutExpired:
            process.kill()
            process.wait(timeout=5)
        shutil.rmtree(root, ignore_errors=True)


def profile():
    return make_runtime_profile(
        profile_id="tenant-product-a4-v2",
        own_bank_holder_aliases=["테스트회사"],
        own_company_aliases=["테스트회사"],
        own_bank_accounts=[
            {"bank_code": "BANKA", "account_number": "111-222-333"},
            {"bank_code": "BANKB", "account_number": "444-555-666"},
        ],
    )


def frame() -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "명세계좌번호": "111-222-333",
                "명세은행코드": "BANKA",
                "거래일자": "2026-07-19",
                "출금": "100,000",
                "입금": "",
                "상대계좌번호": "444-555-666",
                "거래참조번호": "REF-001",
                "적요": "producer-memory-only",
            },
            {
                "명세계좌번호": "444-555-666",
                "명세은행코드": "BANKB",
                "거래일자": "2026-07-19",
                "출금": "",
                "입금": "100,000",
                "상대계좌번호": "111-222-333",
                "거래참조번호": "REF-001",
                "적요": "producer-memory-only",
            },
        ]
    )


def owner_request() -> dict[str, object]:
    closure = {"source_file_sha256": SOURCE_SHA, "source_file_size": SOURCE_SIZE}
    return {
        "run_id": RUN_ID,
        "nonce": "12" * 32,
        "tenant_id": tenant_uuid(profile().profile_id),
        "code_tree_oid": TREE,
        "source_file_sha256": SOURCE_SHA,
        "source_file_size": SOURCE_SIZE,
        "input_closure_digest": digest_object(closure),
    }


def compile_once(store: SQLiteAssignmentStore) -> CompileReceipt:
    request = owner_request()
    registry = store.replace_a4_own_account_registry(
        tenant_id=str(request["tenant_id"]),
        accounts=profile().own_bank_accounts,
        approval_id="a" * 64,
        observed_at=NOW.isoformat().replace("+00:00", "Z"),
    )
    return compile_dataframe(
        frame=frame(),
        tenant_id=str(request["tenant_id"]),
        run_id=RUN_ID,
        source_file_sha256=SOURCE_SHA,
        source_file_size=SOURCE_SIZE,
        adapter=approved_split_adapter(),
        token_service=store.tokens,
        own_account_registry=registry,
    )


def policy_files(
    tmp_path: Path,
    compiled: CompileReceipt,
    *,
    fee_rules=None,
    request_override: dict[str, object] | None = None,
):
    private = Ed25519PrivateKey.generate()
    public = private.public_key().public_bytes_raw()
    trust = {
        "schema_version": "1.0.0",
        "enabled": True,
        "policy_type": "A4_RECON_POLICY",
        "algorithm": "Ed25519",
        "expected_signer_key_id": "finance-test-2026",
        "public_key_b64": base64.b64encode(public).decode("ascii"),
        "revocation_epoch": 7,
    }
    payload = {
        "min_signed_gap_days": 0,
        "max_signed_gap_days": 1,
        "max_component_nodes": 20,
        "max_candidate_edges": 10_000,
        "max_transactions": 100_000,
        "max_pair_checks": 100_000,
        "max_trace_bucket": 1_000,
        "max_memory_bytes": 536_870_912,
        "fee_rules": fee_rules or [],
    }
    request = request_override or owner_request()
    receipt = {
        "policy_type": "A4_RECON_POLICY",
        "policy_id": "a4-policy-test",
        "policy_version": "2.0.0",
        "policy_digest": digest_object(payload),
        "finance_approval_id": "finance-approval-test",
        "finance_approver_role": "FINANCE_APPROVER",
        "signer_key_id": trust["expected_signer_key_id"],
        "issued_at": "2026-07-01T00:00:00Z",
        "effective_from": "2026-07-01T00:00:00Z",
        "expires_at": "2026-12-31T23:59:59Z",
        "revocation_epoch": 7,
        "tenant_id": request["tenant_id"],
        "schema_digest": SCHEMA_DIGEST,
        "adapter_digest": compiled.adapter_digest,
        "account_snapshot_digest": compiled.account_snapshot_digest,
        "dictionary_digest": compiled.dictionary_digest,
        "registry_digest": compiled.registry_digest,
        "code_tree_oid": TREE,
        "run_nonce_digest": sha256_bytes(bytes.fromhex(str(request["nonce"]))),
    }
    signature = private.sign(jcs_bytes({"receipt": receipt, "payload": payload}))
    document = {
        "receipt": {
            **receipt,
            "signature_b64": base64.b64encode(signature).decode("ascii"),
        },
        "payload": payload,
    }
    trust_path, policy_path = tmp_path / "trust.json", tmp_path / "policy.json"
    trust_path.write_bytes(jcs_bytes(trust))
    policy_path.write_bytes(jcs_bytes(document))
    return policy_path, trust_path, document


def make_store(tmp_path: Path) -> SQLiteAssignmentStore:
    return SQLiteAssignmentStore(
        tmp_path / "canonical.sqlite3", TokenService(MemoryKeyStore(key=b"k" * 32))
    )


def registered_registry(store: SQLiteAssignmentStore, request: dict[str, object]):
    return store.replace_a4_own_account_registry(
        tenant_id=str(request["tenant_id"]),
        accounts=profile().own_bank_accounts,
        approval_id="a" * 64,
        observed_at=NOW.isoformat().replace("+00:00", "Z"),
    )


def snapshot_fixture(tmp_path: Path, store: SQLiteAssignmentStore):
    source = tmp_path / "bank.csv"
    frame().to_csv(source, index=False, encoding="utf-8-sig")
    source.chmod(0o600)
    snapshot = secure_source_snapshot(source)
    request = {
        **owner_request(),
        "source_file_sha256": snapshot.receipt["source_file_sha256"],
        "source_file_size": snapshot.receipt["source_file_size"],
    }
    request["input_closure_digest"] = digest_object(
        {
            "source_file_sha256": request["source_file_sha256"],
            "source_file_size": request["source_file_size"],
        }
    )
    from butler_pc_core.accounting.classifier import read_source_frame_fd

    compiled = compile_dataframe(
        frame=read_source_frame_fd(snapshot.producer_fd, snapshot.suffix),
        tenant_id=str(request["tenant_id"]),
        run_id=RUN_ID,
        source_file_sha256=str(request["source_file_sha256"]),
        source_file_size=int(request["source_file_size"]),
        adapter=approved_split_adapter(),
        token_service=store.tokens,
        own_account_registry=registered_registry(store, request),
    )
    return source, snapshot, request, compiled


def test_strict_json_and_integer_only_jcs_reject_ambiguous_input():
    with pytest.raises(A4ContractError, match="BLOCK_DUPLICATE_KEY"):
        strict_json_loads('{"a":1,"a":2}')
    with pytest.raises(A4ContractError, match="BLOCK_NUMBER"):
        strict_json_loads('{"a":NaN}')
    with pytest.raises(A4ContractError, match="BLOCK_NUMBER"):
        strict_json_loads('{"a":1.5}')
    with pytest.raises(A4ContractError, match="BLOCK_NUMBER"):
        jcs_bytes({"money": 1.5})


def test_compiler_uses_stable_uuid_source_identity_and_asia_seoul(tmp_path):
    store = make_store(tmp_path)
    first = compile_once(store)
    second = compile_dataframe(
        frame=frame().copy(),
        tenant_id=owner_request()["tenant_id"],
        run_id=RUN_ID,
        source_file_sha256=SOURCE_SHA,
        source_file_size=SOURCE_SIZE,
        adapter=approved_split_adapter(),
        token_service=store.tokens,
        own_account_registry=store.a4_own_account_registry(owner_request()["tenant_id"]),
    )
    assert first.account_snapshot_digest == second.account_snapshot_digest
    assert all(uuid.UUID(item.txn_uid) for item in first.transactions)
    assert all(len(item.artifact_token) == 32 for item in first.transactions)
    assert {item.booked_at_utc.hour for item in first.transactions} == {15}
    assert len({item.graph_key for item in first.transactions}) == 2


def test_compiler_blocks_extra_amount_column_datetime_and_partial_output(tmp_path):
    store = make_store(tmp_path)
    compile_once(store)
    registry = store.a4_own_account_registry(owner_request()["tenant_id"])
    ambiguous = frame().assign(금액="100000")
    with pytest.raises(A4ContractError, match="BLOCK_AMBIGUOUS_COLUMN"):
        compile_dataframe(
            frame=ambiguous,
            tenant_id=owner_request()["tenant_id"],
            run_id=RUN_ID,
            source_file_sha256=SOURCE_SHA,
            source_file_size=SOURCE_SIZE,
            adapter=approved_split_adapter(),
            token_service=store.tokens,
            own_account_registry=registry,
        )
    bad_date = frame().astype(object)
    bad_date.loc[1, "거래일자"] = datetime(2026, 7, 19)
    with pytest.raises(A4ContractError, match="BLOCK_TYPE"):
        compile_dataframe(
            frame=bad_date,
            tenant_id=owner_request()["tenant_id"],
            run_id=RUN_ID,
            source_file_sha256=SOURCE_SHA,
            source_file_size=SOURCE_SIZE,
            adapter=approved_split_adapter(),
            token_service=store.tokens,
            own_account_registry=registry,
        )
    unregistered = frame()
    unregistered.loc[0, "명세계좌번호"] = "999-888-777"
    with pytest.raises(A4ContractError, match="BLOCK_REGISTRY_IDENTITY"):
        compile_dataframe(
            frame=unregistered,
            tenant_id=owner_request()["tenant_id"],
            run_id=RUN_ID,
            source_file_sha256=SOURCE_SHA,
            source_file_size=SOURCE_SIZE,
            adapter=approved_split_adapter(),
            token_service=store.tokens,
            own_account_registry=registry,
        )
    wrong_bank = frame()
    wrong_bank.loc[0, "명세은행코드"] = "BANKB"
    with pytest.raises(A4ContractError, match="BLOCK_REGISTRY_IDENTITY"):
        compile_dataframe(
            frame=wrong_bank,
            tenant_id=owner_request()["tenant_id"],
            run_id=RUN_ID,
            source_file_sha256=SOURCE_SHA,
            source_file_size=SOURCE_SIZE,
            adapter=approved_split_adapter(),
            token_service=store.tokens,
            own_account_registry=registry,
        )


def direct_policy(*, fee_rules=()) -> VerifiedPolicy:
    return VerifiedPolicy({}, {}, 0, 1, 20, 10_000, tuple(fee_rules))


def tx(
    uid: int,
    direction: Direction,
    account: int,
    *,
    at: int = 0,
    amount: int = 10_000,
    ref: str | None = None,
    kind: str = "PRINCIPAL",
    counterparty: int | None = -1,
    resolution: CounterpartyResolution | None = None,
) -> CompiledTransaction:
    tenant = "00000000-0000-4000-8000-000000000001"
    txn_uid = str(uuid.UUID(int=uid + 100))
    if counterparty == -1:
        counterparty = 2 if account == 1 else 1
    if resolution is None:
        resolution = (
            CounterpartyResolution.EXACT_VERIFIED
            if counterparty is not None
            else CounterpartyResolution.MISSING
        )
    resolved_counterparty = (
        str(uuid.UUID(int=counterparty + 10))
        if resolution is CounterpartyResolution.EXACT_VERIFIED
        and counterparty is not None
        else None
    )
    return CompiledTransaction(
        tenant,
        txn_uid,
        hashlib.sha256(f"row-{uid}".encode()).hexdigest(),
        uid,
        hashlib.sha256(f"canonical-{uid}".encode()).hexdigest(),
        str(uuid.UUID(int=account + 10)),
        direction,
        amount,
        "KRW",
        NOW + timedelta(seconds=at),
        NOW.date(),
        f"{uid + 1:032x}",
        f"BANK{account}",
        ref,
        None,
        resolved_counterparty,
        resolution,
        kind,
        "DEFAULT",
        "DEFAULT",
    )


def test_matcher_2x2_tie_and_unique_optimum_extra_edge_are_correct():
    tied = (
        tx(1, Direction.OUTFLOW, 1),
        tx(2, Direction.OUTFLOW, 1),
        tx(3, Direction.INFLOW, 2),
        tx(4, Direction.INFLOW, 2),
    )
    tie_result = reconcile(tied, direct_policy())
    assert tie_result.forced_edges == ()
    assert tie_result.components[0].has_alternative_optimum is True
    assert len(tie_result.components[0].ambiguous_txn_uids) == 4
    unique = (
        tx(11, Direction.OUTFLOW, 1, at=0),
        tx(12, Direction.OUTFLOW, 1, at=10),
        tx(13, Direction.INFLOW, 2, at=0),
        tx(14, Direction.INFLOW, 2, at=10),
    )
    unique_result = reconcile(unique, direct_policy())
    assert len(unique_result.forced_edges) == 2
    assert unique_result.components[0].has_alternative_optimum is False
    assert unique_result.components[0].ambiguous_txn_uids == ()
    assert unique_result.candidate_pair_checks == 4
    assert unique_result.max_candidate_bucket_size == 2


def test_candidate_requires_reciprocal_stable_account_binding():
    same_amount_date_only = (
        tx(101, Direction.OUTFLOW, 1, counterparty=None),
        tx(102, Direction.INFLOW, 2, counterparty=None),
    )
    assert reconcile(same_amount_date_only, direct_policy()).edges == ()

    one_sided = (
        tx(103, Direction.OUTFLOW, 1, counterparty=2),
        tx(104, Direction.INFLOW, 2, counterparty=None),
    )
    assert reconcile(one_sided, direct_policy()).edges == ()

    reference_only = (
        tx(105, Direction.OUTFLOW, 1, ref="trace", counterparty=None),
        tx(106, Direction.INFLOW, 2, ref="trace", counterparty=None),
    )
    assert reconcile(reference_only, direct_policy()).edges == ()

    reciprocal = (
        tx(107, Direction.OUTFLOW, 1, counterparty=2),
        tx(108, Direction.INFLOW, 2, counterparty=1),
    )
    result = reconcile(reciprocal, direct_policy())
    assert len(result.edges) == 1
    assert result.edges[0].eligibility_basis is EligibilityBasis.RECIPROCAL_ACCOUNT


def test_candidate_builder_blocks_before_partial_result_on_signed_resource_caps():
    rows = (
        tx(111, Direction.OUTFLOW, 1),
        tx(112, Direction.INFLOW, 2),
        tx(113, Direction.INFLOW, 2),
    )
    policy = VerifiedPolicy({}, {}, 0, 1, 20, 10_000, (), 100, 1, 10, 1_048_576)
    with pytest.raises(A4ContractError, match="BLOCK_RESOURCE_PAIR_CHECKS"):
        reconcile(rows, policy)


def test_fee_is_never_inferred_without_one_exact_signed_rule():
    rows = (
        tx(21, Direction.OUTFLOW, 1, amount=10_500),
        tx(22, Direction.INFLOW, 2, amount=10_000),
    )
    assert reconcile(rows, direct_policy()).edges == ()
    rule = FeeRule(
        "fee-500",
        "BANK1",
        "DEFAULT",
        "DEFAULT",
        1,
        100_000,
        500,
        NOW - timedelta(days=1),
        NOW + timedelta(days=1),
    )
    result = reconcile(rows, direct_policy(fee_rules=(rule,)))
    assert result.edges[0].fee_mode is FeeMode.EMBEDDED_DEBIT
    assert (
        result.edges[0].evidence_grade is EvidenceGrade.SCHEDULE_SUPPORTED_FEE_CANDIDATE
    )
    with pytest.raises(A4ContractError, match="BLOCK_AMBIGUOUS_FEE_RULE"):
        reconcile(rows, direct_policy(fee_rules=(rule, rule)))


def test_explicit_fee_row_replaces_no_fee_candidate_and_is_one_to_one():
    rows = (
        tx(31, Direction.OUTFLOW, 1, amount=10_000, ref="same"),
        tx(32, Direction.INFLOW, 2, amount=10_000, ref="same"),
        tx(
            33,
            Direction.OUTFLOW,
            1,
            amount=500,
            ref="same",
            kind="FEE",
        ),
    )
    result = reconcile(rows, direct_policy())
    assert len(result.edges) == 1
    assert result.edges[0].fee_mode is FeeMode.SEPARATE_FEE_ROW
    assert result.edges[0].fee is not None
    assert result.edges[0].fee.txn_uid == rows[2].txn_uid


def test_signed_policy_rejects_wrong_binding_and_expiry(tmp_path):
    store = make_store(tmp_path)
    compiled = compile_once(store)
    policy_path, trust_path, _ = policy_files(tmp_path, compiled)
    trust_doc = strict_json_loads(trust_path.read_bytes())
    trust = TrustPolicy(
        trust_doc["expected_signer_key_id"],
        trust_doc["public_key_b64"],
        trust_doc["revocation_epoch"],
        ("A4_RECON_POLICY",),
        "Ed25519",
    )
    request = owner_request()
    expected = {
        "tenant_id": request["tenant_id"],
        "schema_digest": SCHEMA_DIGEST,
        "adapter_digest": compiled.adapter_digest,
        "account_snapshot_digest": compiled.account_snapshot_digest,
        "dictionary_digest": compiled.dictionary_digest,
        "registry_digest": compiled.registry_digest,
        "code_tree_oid": TREE,
        "run_nonce_digest": sha256_bytes(bytes.fromhex(request["nonce"])),
    }
    assert verify_signed_policy(
        policy_path.read_bytes(), trust=trust, expected=expected, now=NOW
    )
    with pytest.raises(A4ContractError, match="BLOCK_BINDING"):
        verify_signed_policy(
            policy_path.read_bytes(),
            trust=trust,
            expected={**expected, "code_tree_oid": "b" * 40},
            now=NOW,
        )
    with pytest.raises(A4ContractError, match="BLOCK_EXPIRED"):
        verify_signed_policy(
            policy_path.read_bytes(),
            trust=trust,
            expected=expected,
            now=datetime(2027, 1, 1, tzinfo=timezone.utc),
        )


def test_product_db_outbox_export_and_independent_verifier(tmp_path, monkeypatch):
    store = make_store(tmp_path)
    _source, snapshot, request, compiled = snapshot_fixture(tmp_path, store)
    policy_path, trust_path, _ = policy_files(
        tmp_path, compiled, request_override=request
    )
    monkeypatch.setattr(service_v2, "PINNED_TRUST_POLICY_PATH", trust_path)
    evidence_root = tmp_path / "evidence"
    result = run_canonical_product_reconciliation(
        frame=None,
        company_profile=profile(),
        token_service=store.tokens,
        store=store,
        owner_request=request,
        policy_path=policy_path,
        trust_policy_path=trust_path,
        evidence_root=evidence_root,
        observed_at=NOW,
        verify_after_export=True,
        source_snapshot=snapshot,
    )
    snapshot.close()
    assert result["state"] == "PUBLISHED"
    assert result["forced_candidate_count"] == 1
    final = evidence_root / RUN_ID
    assert {item.name for item in final.iterdir()} == {
        "run_request.json",
        "input_receipt.json",
        "policy_receipt.json",
        "account_snapshot_receipt.json",
        "graph_observation.json",
        "classification_receipt.json",
        "runtime_invariants.json",
        "dlp_receipt.json",
        "manifest.json",
        "SHA256SUMS",
    }
    combined = b"".join(item.read_bytes() for item in final.iterdir())
    for raw in (b"111-222-333", b"444-555-666", b"producer-memory-only", b"REF-001"):
        assert raw not in combined
    assert (
        store.reconciliation_run(request["tenant_id"], RUN_ID)["state"]
        == "PUBLISHED"
    )
    with store._connect() as conn:
        assert (
            conn.execute(
                "SELECT COUNT(*) FROM recon_outbox WHERE state='STAGED'"
            ).fetchone()[0]
            == 10
        )
        assert conn.execute("SELECT COUNT(*) FROM recon_review").fetchone()[0] == 0
        stored_receipt = strict_json_loads(
            bytes(
                conn.execute(
                    "SELECT receipt_jcs FROM recon_verification_receipt"
                ).fetchone()[0]
            )
        )
        assert stored_receipt["authority_id"] == "butler-a4-isolated-test-authority"
        assert stored_receipt["signer_revocation_epoch"] == 9
        assert conn.execute("PRAGMA integrity_check").fetchone()[0] == "ok"
    forged_body = {key: value for key, value in stored_receipt.items() if key != "signature"}
    forged_body["signer_revocation_epoch"] = 10
    forged_signature = Ed25519PrivateKey.generate().sign(jcs_bytes(forged_body))
    forged_receipt = {
        **forged_body,
        "signature": base64.urlsafe_b64encode(forged_signature)
        .decode("ascii")
        .rstrip("="),
    }
    with pytest.raises(AssignmentError, match="BLOCK_RECEIPT_BINDING"):
        store.mark_reconciliation_verified(
            str(request["tenant_id"]), RUN_ID, forged_receipt
        )


def test_actual_a4_product_endpoints_read_candidates_and_idempotent_review(
    tmp_path, monkeypatch
):
    current_profile = profile()
    runtime = AccountingReviewRuntime(
        db_path=tmp_path / "endpoint.sqlite3",
        key_store=MemoryKeyStore(key=b"p" * 32),
    )
    _source, snapshot, request, compiled = snapshot_fixture(tmp_path, runtime.store)
    policy_path, trust_path, _ = policy_files(
        tmp_path, compiled, request_override=request
    )
    monkeypatch.setattr(service_v2, "PINNED_TRUST_POLICY_PATH", trust_path)
    run_canonical_product_reconciliation(
        frame=None,
        company_profile=current_profile,
        token_service=runtime.tokens,
        store=runtime.store,
        owner_request=request,
        policy_path=policy_path,
        trust_policy_path=trust_path,
        evidence_root=tmp_path / "endpoint-evidence",
        observed_at=NOW,
        verify_after_export=True,
        source_snapshot=snapshot,
    )
    snapshot.close()

    class ProfileStore:
        def load_active_profile(self):
            return current_profile

    token_manager = CapabilityTokenManager(tmp_path / "endpoint-token")
    token = token_manager.generate()
    monkeypatch.setattr(assignment_routes, "CompanyProfileStore", ProfileStore)
    monkeypatch.setattr(assignment_routes, "_TOKEN_MANAGER", token_manager)
    set_accounting_review_runtime_for_tests(runtime)
    app = FastAPI()
    app.include_router(assignment_routes.router)
    client = TestClient(app, headers={"Authorization": f"Bearer {token}"})
    try:
        created = client.post("/box5/reconciliation/a4/runs", json={"run_id": RUN_ID})
        assert created.status_code == 200
        assert created.json()["state"] == "PUBLISHED"
        candidates = client.get(f"/box5/reconciliation/a4/runs/{RUN_ID}/candidates")
        assert candidates.status_code == 200
        assert candidates.json()["affects_reporting"] is False
        edge_id = candidates.json()["items"][0]["edge_id"]
        path = f"/box5/reconciliation/a4/runs/{RUN_ID}/candidates/{edge_id}/review"
        first = client.post(
            path,
            headers={"Idempotency-Key": "a4-review-endpoint-0001"},
            json={"decision": "DEFER"},
        )
        replay = client.post(
            path,
            headers={"Idempotency-Key": "a4-review-endpoint-0001"},
            json={"decision": "DEFER"},
        )
        assert first.status_code == replay.status_code == 200
        assert first.json()["affects_reporting"] is False
        assert replay.json()["replayed"] is True
    finally:
        set_accounting_review_runtime_for_tests(None)


def test_v51_published_candidates_and_review_are_denied_after_trust_revocation(
    tmp_path, monkeypatch, isolated_verifier_authority
):
    store = make_store(tmp_path)
    _source, snapshot, request, compiled = snapshot_fixture(tmp_path, store)
    policy_path, finance_trust_path, _ = policy_files(
        tmp_path, compiled, request_override=request
    )
    monkeypatch.setattr(service_v2, "PINNED_TRUST_POLICY_PATH", finance_trust_path)
    try:
        run_canonical_product_reconciliation(
            frame=None,
            company_profile=profile(),
            token_service=store.tokens,
            store=store,
            owner_request=request,
            policy_path=policy_path,
            trust_policy_path=finance_trust_path,
            evidence_root=tmp_path / "revocation-evidence",
            observed_at=NOW,
            verify_after_export=True,
            source_snapshot=snapshot,
        )
    finally:
        snapshot.close()

    candidates = store.reconciliation_candidates(str(request["tenant_id"]), RUN_ID)
    assert candidates
    revoked_trust = {
        **isolated_verifier_authority["trust"],
        "enabled": False,
    }
    revoked_path = tmp_path / "revoked-authority-trust.json"
    revoked_path.write_bytes(jcs_bytes(revoked_trust))
    revoked_path.chmod(0o600)
    store._a4_authority_trust_path = revoked_path

    with pytest.raises(AssignmentError, match="BLOCK_UNVERIFIED_ACCESS"):
        store.reconciliation_candidates(str(request["tenant_id"]), RUN_ID)
    with pytest.raises(AssignmentError, match="BLOCK_UNVERIFIED_ACCESS"):
        store.append_reconciliation_review(
            tenant_id=str(request["tenant_id"]),
            run_id=RUN_ID,
            edge_id=candidates[0]["edge_id"],
            decision="DEFER",
            actor_id="reviewer",
            idempotency_key="post-revocation-review-attempt",
        )


def test_export_retry_reverifies_full_file_set_and_tamper_blocks(tmp_path, monkeypatch):
    store = make_store(tmp_path)
    _source, snapshot, request, compiled = snapshot_fixture(tmp_path, store)
    policy_path, trust_path, _ = policy_files(
        tmp_path, compiled, request_override=request
    )
    monkeypatch.setattr(service_v2, "PINNED_TRUST_POLICY_PATH", trust_path)
    root = tmp_path / "evidence"
    kwargs = dict(
        frame=None,
        company_profile=profile(),
        token_service=store.tokens,
        store=store,
        owner_request=request,
        policy_path=policy_path,
        trust_policy_path=trust_path,
        evidence_root=root,
        observed_at=NOW,
        verify_after_export=True,
        source_snapshot=snapshot,
    )
    run_canonical_product_reconciliation(**kwargs)
    run_canonical_product_reconciliation(**kwargs)
    target = root / RUN_ID / "classification_receipt.json"
    target.write_bytes(b"{}")
    with pytest.raises(A4ContractError, match="BLOCK_IDEMPOTENCY_CONFLICT"):
        run_canonical_product_reconciliation(**kwargs)
    snapshot.close()
    assert any(item.name.startswith("quarantine-") for item in root.iterdir())


def _rehash_evidence_bundle(root: Path) -> None:
    manifest = strict_json_loads((root / "manifest.json").read_bytes())
    json_names = sorted(service_v2.EVIDENCE_FILES - {"manifest.json", "SHA256SUMS"})
    manifest["files"] = [
        {
            "path": name,
            "sha256": sha256_bytes((root / name).read_bytes()),
            "size": len((root / name).read_bytes()),
        }
        for name in json_names
    ]
    (root / "manifest.json").write_bytes(jcs_bytes(manifest))
    checksums = b"".join(
        f"{sha256_bytes((root / name).read_bytes())}  {name}\n".encode("ascii")
        for name in sorted(service_v2.EVIDENCE_FILES - {"SHA256SUMS"})
    )
    (root / "SHA256SUMS").write_bytes(checksums)


def _physical_row_root(rows: list[dict[str, object]]) -> str:
    ordered = hashlib.sha256()
    for row in rows:
        for name in ("row_locator", "cell_vector_digest", "disposition"):
            encoded = str(row[name]).encode("utf-8")
            ordered.update(len(encoded).to_bytes(4, "big"))
            ordered.update(encoded)
    return ordered.hexdigest()


def test_independent_verifier_blocks_omitted_extra_and_changed_binding_edges(
    tmp_path, monkeypatch
):
    store = make_store(tmp_path)
    _source, snapshot, request, compiled = snapshot_fixture(tmp_path, store)
    policy_path, trust_path, _ = policy_files(
        tmp_path, compiled, request_override=request
    )
    monkeypatch.setattr(service_v2, "PINNED_TRUST_POLICY_PATH", trust_path)
    evidence_root = tmp_path / "evidence"
    run_canonical_product_reconciliation(
        frame=None,
        company_profile=profile(),
        token_service=store.tokens,
        store=store,
        owner_request=request,
        policy_path=policy_path,
        trust_policy_path=trust_path,
        evidence_root=evidence_root,
        observed_at=NOW,
        verify_after_export=True,
        source_snapshot=snapshot,
    )
    original = evidence_root / RUN_ID

    for attack in (
        "OMIT",
        "EXTRA",
        "BINDING",
        "FORGED_PRODUCT_CODE_NOT_IN_SOURCE",
        "ROW_INJECTION",
        "CELL_VALUE_SWAP",
    ):
        attacked = tmp_path / f"attack-{attack.lower()}"
        shutil.copytree(original, attacked)
        graph_path = attacked / "graph_observation.json"
        graph = strict_json_loads(graph_path.read_bytes())
        if attack == "OMIT":
            graph["edges"] = []
        elif attack == "EXTRA":
            extra = {**graph["edges"][0], "edge_id": "f" * 64}
            graph["edges"].append(extra)
        elif attack == "BINDING":
            graph["edges"][0]["eligibility_basis"] = "TRACE_AND_ACCOUNT"
        elif attack == "FORGED_PRODUCT_CODE_NOT_IN_SOURCE":
            input_path = attacked / "input_receipt.json"
            input_receipt = strict_json_loads(input_path.read_bytes())
            input_receipt["compiled_transactions"][0]["product_code_id"] = "TRANSFER"
            input_path.write_bytes(jcs_bytes(input_receipt))
        elif attack == "ROW_INJECTION":
            input_path = attacked / "input_receipt.json"
            input_receipt = strict_json_loads(input_path.read_bytes())
            closure = input_receipt["row_closure"]
            injected = dict(closure["rows"][-1])
            injected["row_ordinal"] = closure["row_count"]
            injected["row_locator"] = "f" * 64
            injected["cell_vector_digest"] = "e" * 64
            injected["disposition"] = "BLANK"
            injected["transaction_ordinal"] = None
            closure["rows"].append(injected)
            closure["row_count"] += 1
            closure["disposition_counts"]["BLANK"] += 1
            closure["ordered_row_root"] = _physical_row_root(closure["rows"])
            input_path.write_bytes(jcs_bytes(input_receipt))
        elif attack == "CELL_VALUE_SWAP":
            input_path = attacked / "input_receipt.json"
            input_receipt = strict_json_loads(input_path.read_bytes())
            rows = input_receipt["row_closure"]["rows"]
            rows[1]["cell_vector_digest"], rows[2]["cell_vector_digest"] = (
                rows[2]["cell_vector_digest"],
                rows[1]["cell_vector_digest"],
            )
            input_receipt["row_closure"]["ordered_row_root"] = _physical_row_root(
                rows
            )
            input_path.write_bytes(jcs_bytes(input_receipt))
        if attack in {"OMIT", "EXTRA", "BINDING"}:
            graph_path.write_bytes(jcs_bytes(graph))
        _rehash_evidence_bundle(attacked)
        with pytest.raises(A4ContractError, match="BLOCK_OFFLINE_VERIFIER"):
            service_v2.verify_in_separate_process(
                evidence_dir=attacked,
                trust_policy_path=trust_path,
                token_service=store.tokens,
                tenant_id=str(request["tenant_id"]),
                source_fd=snapshot.verifier_fd,
                source_suffix=snapshot.suffix,
            )
    snapshot.close()


def test_value_dlp_detects_allowed_looking_pii_and_paths():
    with pytest.raises(A4ContractError, match="BLOCK_DLP"):
        closed_world_dlp({"receipt.json": {"safe_value": "010-1234-5678"}})
    with pytest.raises(A4ContractError, match="BLOCK_DLP"):
        closed_world_dlp({"receipt.json": {"safe_value": "/Users/alice/file"}})
    encoded_account = base64.urlsafe_b64encode(b"010-1234-5678").decode("ascii")
    with pytest.raises(A4ContractError, match="BLOCK_DLP"):
        closed_world_dlp({"receipt.json": {"safe_value": encoded_account}})
    with pytest.raises(A4ContractError, match="BLOCK_DLP"):
        closed_world_dlp(
            {"receipt.json": {"safe_value": b"user@example.com".hex()}}
        )


def test_v51_completion_status_has_exact_findings_and_remains_disabled():
    root = Path(__file__).resolve().parents[3]
    schema = strict_json_loads(
        (
            root
            / "butler_pc_core/accounting/classify/contracts/a4_v31"
            / "completion_status.schema.json"
        ).read_bytes()
    )
    status = strict_json_loads(
        (root / "docs/ops/A4_COMPLETION_STATUS.json").read_bytes()
    )
    Draft202012Validator.check_schema(schema)
    Draft202012Validator(schema).validate(status)
    finding_ids = [item["id"] for item in status["findings"]]
    assert len(finding_ids) == len(set(finding_ids)) == 36
    v51 = [item for item in status["findings"] if item["id"].startswith("A4V51-")]
    assert len(v51) == 5
    assert len({item["evidence_digest"] for item in v51}) == 5
    local_evidence = strict_json_loads(
        (root / "docs/ops/evidence/A4_V51_LOCAL_EVIDENCE.json").read_bytes()
    )
    evidence_by_id = {
        item["finding_id"]: item for item in local_evidence["findings"]
    }
    assert set(evidence_by_id) == {item["id"] for item in v51}
    for finding in v51:
        evidence = evidence_by_id[finding["id"]]
        product_file = root / evidence["primary_product_file"]
        assert finding["status"] == evidence["status"]
        assert finding["evidence_digest"] == evidence["evidence_digest"]
        assert hashlib.sha256(product_file.read_bytes()).hexdigest() == evidence[
            "evidence_digest"
        ]
    assert status["code_pass"] is False
    assert status["runtime_activation_allowed"] is False
    release = strict_json_loads(
        (
            root
            / "butler_pc_core/accounting/classify/contracts/a4_v31"
            / "release_manifest.production.json"
        ).read_bytes()
    )
    assert release == {
        "schema_version": "butler.box5.a4.release_manifest.v1",
        "product_version": "5.0.0",
        "contract_version": "3.2.0",
        "canonical_base_commit_oid": "b38e9e3321c0b972fb257ebd3831fa472da9c5a4",
        "database_migration": "a4_store_schema_v32",
        "state_machine": "recon_run_v3.2",
        "verifier": "BUTLER_A4_ISOLATED_RAW_COMPILER_V3.2",
        "verifier_authority": "unix-scm-rights-v1",
        "verification_receipt": "butler.box5.a4.verification_receipt.v3.2",
        "code_dictionary": "BUTLER_A4_KR_CODES@3.1.0",
        "runtime_activation_allowed": False,
    }


def test_v5_producer_has_no_verifier_private_key_or_signing_api(
    isolated_verifier_authority,
):
    service = TokenService(MemoryKeyStore(key=b"K" * 32))
    material = service.a4_verifier_material(
        "11111111-1111-4111-8111-111111111111"
    )
    assert set(material) == {
        "schema_version",
        "tenant_id",
        "key_id",
        "own_account_key_b64",
        "bank_reference_key_b64",
        "counterparty_account_key_b64",
        "run_transaction_key_b64",
    }
    assert "verification_signing_seed_b64" not in material
    assert not hasattr(service, "a4_verifier_public_key")
    with pytest.raises(A4ContractError, match="BLOCK_TRUST_UNAVAILABLE"):
        load_authority_trust(
            isolated_verifier_authority["trust_path"],
            now=NOW,
            allow_test_authority=False,
        )


def test_v51_stalled_authority_client_does_not_block_next_client(
    tmp_path, isolated_verifier_authority
):
    stalled = socket.socket(socket.AF_UNIX, socket.SOCK_STREAM)
    stalled.settimeout(1)
    stalled.connect(str(isolated_verifier_authority["socket_path"]))
    source = tmp_path / "concurrency-source.csv"
    source.write_bytes(b"header\nvalue\n")
    source.chmod(0o600)
    source_fd = os.open(source, os.O_RDONLY)
    started = time.monotonic()
    try:
        with socket.socket(socket.AF_UNIX, socket.SOCK_STREAM) as next_client:
            next_client.settimeout(1.5)
            next_client.connect(str(isolated_verifier_authority["socket_path"]))
            _send_request(next_client, jcs_bytes({}), source_fd)
            response = _receive_response(next_client)
        assert response["status"] == "BLOCK"
        assert time.monotonic() - started < 1.5
    finally:
        os.close(source_fd)
        stalled.close()


def test_v51_authority_applies_bounded_backpressure(isolated_verifier_authority):
    clients: list[socket.socket] = []
    try:
        for _ in range(AUTHORITY_MAX_CLIENTS):
            client = socket.socket(socket.AF_UNIX, socket.SOCK_STREAM)
            client.settimeout(1)
            client.connect(str(isolated_verifier_authority["socket_path"]))
            clients.append(client)
        time.sleep(0.05)
        with socket.socket(socket.AF_UNIX, socket.SOCK_STREAM) as overflow:
            overflow.settimeout(1)
            overflow.connect(str(isolated_verifier_authority["socket_path"]))
            assert overflow.recv(1) == b""
    finally:
        for client in clients:
            client.close()
        time.sleep(0.05)


def test_v5_bundled_unapproved_authority_is_fail_closed():
    contract_root = (
        Path(__file__).resolve().parents[3]
        / "butler_pc_core/accounting/classify/contracts/a4_v31"
    )
    schema = strict_json_loads(
        (contract_root / "verifier_authority_trust.schema.json").read_bytes()
    )
    bundled = strict_json_loads(
        (contract_root / "verifier_authority_trust.production.json").read_bytes()
    )
    Draft202012Validator.check_schema(schema)
    Draft202012Validator(schema).validate(bundled)
    assert bundled["enabled"] is False
    with pytest.raises(A4ContractError, match="BLOCK_TRUST_UNAVAILABLE"):
        load_authority_trust(
            contract_root / "verifier_authority_trust.production.json",
            now=NOW,
        )


@pytest.mark.parametrize(
    ("field", "value"),
    [
        ("enabled", False),
        ("not_after_utc", "2026-01-01T00:00:00Z"),
        ("revocation_epoch", -1),
    ],
)
def test_v5_disabled_expired_or_revoked_authority_is_rejected(
    tmp_path, isolated_verifier_authority, field, value
):
    trust = {**isolated_verifier_authority["trust"], field: value}
    trust_path = tmp_path / "invalid-authority-trust.json"
    trust_path.write_bytes(jcs_bytes(trust))
    trust_path.chmod(0o600)
    with pytest.raises(A4ContractError, match="BLOCK_TRUST_UNAVAILABLE"):
        load_authority_trust(
            trust_path,
            now=NOW,
            allow_test_authority=True,
        )


def test_v5_public_key_substitution_does_not_match_running_authority(
    tmp_path, isolated_verifier_authority
):
    substituted_key = Ed25519PrivateKey.generate().public_key().public_bytes_raw()
    substituted_trust = {
        **isolated_verifier_authority["trust"],
        "public_key_b64": base64.b64encode(substituted_key).decode("ascii"),
    }
    trust_path = tmp_path / "substituted-authority-trust.json"
    trust_path.write_bytes(jcs_bytes(substituted_trust))
    trust_path.chmod(0o600)
    source_path = tmp_path / "source.csv"
    source_path.write_bytes(b"header\nvalue\n")
    source_path.chmod(0o600)
    descriptor = os.open(source_path, os.O_RDONLY)
    try:
        with pytest.raises(A4ContractError, match="BLOCK_OFFLINE_VERIFIER"):
            request_authority_verification(
                authority_socket=isolated_verifier_authority["socket_path"],
                authority_trust_path=trust_path,
                evidence_files={"run_request.json": b"{}"},
                finance_trust=b"{}",
                dictionary=b"{}",
                key_material=TokenService(
                    MemoryKeyStore(key=b"s" * 32)
                ).a4_verifier_material(
                    "11111111-1111-4111-8111-111111111111"
                ),
                tenant_id="11111111-1111-4111-8111-111111111111",
                source_fd=descriptor,
                source_suffix=".csv",
                now=NOW,
                request_nonce="f" * 64,
                allow_test_authority=True,
            )
    finally:
        os.close(descriptor)


def test_v5_legacy_signed_receipts_require_explicit_migration(tmp_path):
    store = make_store(tmp_path)
    with store._connect() as conn:
        conn.executescript(
            """
            DROP TRIGGER IF EXISTS recon_review_requires_published_receipt;
            DROP TRIGGER IF EXISTS recon_publish_requires_pass_receipt;
            DROP TRIGGER IF EXISTS recon_verification_receipt_no_update;
            DROP TRIGGER IF EXISTS recon_verification_receipt_no_delete;
            DROP TABLE recon_review;
            DROP TABLE recon_verification_receipt;
            CREATE TABLE recon_verification_receipt (
                tenant_id TEXT NOT NULL,
                run_id TEXT NOT NULL,
                receipt_digest TEXT NOT NULL,
                nonce TEXT NOT NULL,
                evidence_manifest_digest TEXT NOT NULL,
                verifier_id TEXT NOT NULL,
                signer_key_id TEXT NOT NULL,
                code_closure_digest TEXT NOT NULL,
                decision TEXT NOT NULL,
                verified_at TEXT NOT NULL,
                receipt_jcs BLOB NOT NULL,
                PRIMARY KEY (tenant_id, run_id)
            );
            INSERT INTO recon_verification_receipt VALUES (
                'legacy-tenant','legacy-run','legacy-digest','legacy-nonce',
                'legacy-manifest','legacy-verifier','producer-visible-root',
                'legacy-closure','PASS','2026-07-01T00:00:00Z',X'7B7D'
            );
            """
        )
    migrated = SQLiteAssignmentStore(store.path, store.tokens)
    with pytest.raises(
        AssignmentError, match="BLOCK_SCHEMA_MIGRATION_REQUIRED"
    ):
        migrated._require_a4_schema()


@pytest.mark.parametrize("suffix", [".csv", ".xlsx"])
def test_actual_snapshot_fd_product_path_and_independent_raw_replay(
    tmp_path, monkeypatch, suffix
):
    source = tmp_path / f"bank{suffix}"
    if suffix == ".csv":
        frame().to_csv(source, index=False, encoding="utf-8-sig")
    else:
        frame().to_excel(source, index=False)
    source.chmod(0o600)
    snapshot = secure_source_snapshot(source)
    request = {
        **owner_request(),
        "source_file_sha256": snapshot.receipt["source_file_sha256"],
        "source_file_size": snapshot.receipt["source_file_size"],
    }
    request["input_closure_digest"] = digest_object(
        {
            "source_file_sha256": request["source_file_sha256"],
            "source_file_size": request["source_file_size"],
        }
    )
    store = make_store(tmp_path)
    from butler_pc_core.accounting.classifier import read_source_frame_fd

    compiled = compile_dataframe(
        frame=read_source_frame_fd(snapshot.producer_fd, snapshot.suffix),
        tenant_id=str(request["tenant_id"]),
        run_id=RUN_ID,
        source_file_sha256=str(request["source_file_sha256"]),
        source_file_size=int(request["source_file_size"]),
        adapter=approved_split_adapter(),
        token_service=store.tokens,
        own_account_registry=registered_registry(store, request),
    )
    policy_path, trust_path, _ = policy_files(
        tmp_path, compiled, request_override=request
    )
    monkeypatch.setattr(service_v2, "PINNED_TRUST_POLICY_PATH", trust_path)
    source.write_text("replaced-after-owner-boundary", encoding="utf-8")
    try:
        result = run_canonical_product_reconciliation(
            frame=None,
            company_profile=profile(),
            token_service=store.tokens,
            store=store,
            owner_request=request,
            policy_path=policy_path,
            trust_policy_path=trust_path,
            evidence_root=tmp_path / "snapshot-evidence",
            observed_at=NOW,
            verify_after_export=True,
            source_snapshot=snapshot,
        )
    finally:
        snapshot.close()
        assert result["state"] == "PUBLISHED"
    input_receipt = strict_json_loads(
        (tmp_path / "snapshot-evidence" / RUN_ID / "input_receipt.json").read_bytes()
    )
    assert input_receipt["source_snapshot_receipt"]["snapshot_path_unlinked"] is True
    assert input_receipt["row_closure"]["row_count"] == 3
    assert input_receipt["row_closure"]["transaction_row_count"] == 2
    assert input_receipt["row_closure"]["disposition_counts"]["HEADER"] == 1
    assert input_receipt["row_closure"]["disposition_counts"]["PRINCIPAL"] == 2


@pytest.mark.parametrize("unsafe_kind", ["formula", "hidden_sheet"])
def test_actual_xlsx_product_path_rejects_active_content(
    tmp_path, monkeypatch, unsafe_kind
):
    source = tmp_path / f"unsafe-{unsafe_kind}.xlsx"
    frame().to_excel(source, index=False)

    from openpyxl import load_workbook

    workbook = load_workbook(source)
    if unsafe_kind == "formula":
        workbook.active["D2"] = "=50000+50000"
    else:
        workbook.active.sheet_state = "hidden"
        workbook.create_sheet("visible")
    workbook.save(source)
    workbook.close()
    source.chmod(0o600)

    snapshot = secure_source_snapshot(source)
    store = make_store(tmp_path)
    request = {
        **owner_request(),
        "source_file_sha256": snapshot.receipt["source_file_sha256"],
        "source_file_size": snapshot.receipt["source_file_size"],
    }
    request["input_closure_digest"] = digest_object(
        {
            "source_file_sha256": request["source_file_sha256"],
            "source_file_size": request["source_file_size"],
        }
    )
    monkeypatch.setattr(service_v2, "PINNED_TRUST_POLICY_PATH", tmp_path / "unused")
    try:
        with pytest.raises(A4ContractError, match="BLOCK_ADAPTER_POLICY"):
            run_canonical_product_reconciliation(
                frame=None,
                company_profile=profile(),
                token_service=store.tokens,
                store=store,
                owner_request=request,
                policy_path=tmp_path / "unused-policy",
                trust_policy_path=tmp_path / "unused-trust",
                evidence_root=tmp_path / "unsafe-evidence",
                observed_at=NOW,
                verify_after_export=True,
                source_snapshot=snapshot,
            )
    finally:
        snapshot.close()


def test_canonical_run_queue_is_durable_and_events_are_append_only(tmp_path):
    store = make_store(tmp_path)
    request = owner_request()
    queued = queue_canonical_product_reconciliation(
        store=store, owner_request=request, observed_at=NOW
    )
    assert queued["state"] == "QUEUED"
    store.claim_reconciliation_run(
        tenant_id=str(request["tenant_id"]),
        run_id=RUN_ID,
        worker_owner_digest="b" * 64,
        observed_at="2026-07-19T06:00:00Z",
        lease_expires_at="2026-07-19T06:00:01Z",
    )
    assert store.recover_expired_reconciliation_jobs(
        observed_at="2026-07-19T06:00:02Z"
    ) == 1
    assert store.reconciliation_run(str(request["tenant_id"]), RUN_ID)["state"] == "QUEUED"
    with pytest.raises(Exception, match="BLOCK_UNVERIFIED_ACCESS"):
        store.reconciliation_candidates(str(request["tenant_id"]), RUN_ID)
    with pytest.raises(Exception, match="BLOCK_UNVERIFIED_ACCESS"):
        store.append_reconciliation_review(
            tenant_id=str(request["tenant_id"]),
            run_id=RUN_ID,
            edge_id="a" * 64,
            decision="DEFER",
            actor_id="reviewer",
            idempotency_key="unverified-review-attempt",
        )
    with store._connect() as conn, pytest.raises(Exception, match="APPEND_ONLY_A4"):
        conn.execute(
            "UPDATE recon_run_event SET to_state='FAILED' WHERE tenant_id=? AND run_id=?",
            (request["tenant_id"], RUN_ID),
        )


def test_actual_shadow_runner_durably_blocks_unknown_build(tmp_path):
    store = make_store(tmp_path)
    request = owner_request()
    queue_canonical_product_reconciliation(
        store=store, owner_request=request, observed_at=NOW
    )

    class Snapshot:
        closed = False

        def close(self):
            self.closed = True

    snapshot = Snapshot()
    count = run_and_write_product_shadow(
        frame(),
        str(tmp_path / "shadow.jsonl"),
        profile(),
        RUN_ID,
        store.tokens,
        "unknown",
        NOW,
        store,
        request,
        None,
        str(tmp_path / "evidence"),
        snapshot,
    )
    assert count == 0
    assert snapshot.closed is True
    failed = store.reconciliation_run(str(request["tenant_id"]), RUN_ID)
    assert failed["state"] == "FAILED"
    assert failed["error_code"] == "BLOCK_BUILD_IDENTITY_UNKNOWN"
