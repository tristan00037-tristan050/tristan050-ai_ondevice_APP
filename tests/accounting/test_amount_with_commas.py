"""test_amount_with_commas.py — 콤마 포함 금액이 [요약] 시트에 정상 집계돼야 한다."""
from __future__ import annotations

import tempfile
from pathlib import Path

import pytest

try:
    import pandas as pd
    import openpyxl
    _DEPS_OK = True
except ImportError:
    _DEPS_OK = False

_skip = pytest.mark.skipif(not _DEPS_OK, reason="pandas/openpyxl 미설치")

_EXPECTED_TOTAL = 3_200_000 + 88_000 + 1_500_000 + 320_000 + 85_000  # 5,193,000


@_skip
def test_summary_sheet_with_comma_amounts():
    """금액 컬럼이 '1,234,567' 형식일 때 [요약] 시트 합계·총계가 정상값이어야 한다."""
    from butler_pc_core.accounting.classifier import classify_df, save_classified
    from .conftest import make_comma_amount_df

    df = classify_df(make_comma_amount_df())

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        tmp = f.name
    try:
        save_classified(df, tmp)
        wb = openpyxl.load_workbook(tmp)
        ws = wb["요약"]

        total_row = [cell.value for cell in ws[ws.max_row]]
        assert total_row[0] == "총계", f"마지막 행이 총계 행이 아님: {total_row}"

        # [요약] 컬럼: 분류과목 | 구분 | 건수 | 합계금액 | 비율 → 합계금액=index 3
        total_sum = total_row[3]
        assert total_sum > 0, (
            f"합계가 0 — 콤마 포함 금액 파싱 실패. 전체 요약:\n"
            + "\n".join(str([c.value for c in row]) for row in ws.iter_rows())
        )
        assert total_sum == _EXPECTED_TOTAL, (
            f"합계 불일치 — 기대: {_EXPECTED_TOTAL:,}, 실제: {total_sum:,}"
        )
    finally:
        Path(tmp).unlink(missing_ok=True)


@_skip
def test_comma_amounts_do_not_affect_classification():
    """금액 형식(콤마/원화기호)이 계정과목 분류 결과에 영향을 주지 않아야 한다."""
    from butler_pc_core.accounting.classifier import classify_df
    from .conftest import make_comma_amount_df

    result = classify_df(make_comma_amount_df())
    classified = result[result["분류과목"] != "미분류"]

    assert len(classified) >= 4, (
        f"콤마 금액 데이터에서 분류 성공 건수 부족: {len(classified)}/5"
    )
    assert (classified["신뢰도"] >= 0.50).all(), (
        "분류 성공 항목 신뢰도 50% 미만 발견"
    )
