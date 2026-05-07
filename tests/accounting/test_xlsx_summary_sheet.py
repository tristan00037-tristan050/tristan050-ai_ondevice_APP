"""test_xlsx_summary_sheet.py — [요약] 시트 합계금액 컬럼 + 합계 정확성 검증."""
from __future__ import annotations

import tempfile
from pathlib import Path

import pytest

try:
    import pandas as pd
    import openpyxl
    _DEPS_OK = True
except ImportError:
    _DEPS_OK = False

_skip = pytest.mark.skipif(not _DEPS_OK, reason="pandas/openpyxl 미설치")


def _sep_df() -> "pd.DataFrame":
    """입출금 분리 컬럼 DataFrame — _amt 컬럼 생성 경로."""
    return pd.DataFrame({
        "적요":  ["급여 지급",  "통신비 납부", "급여 지급"],
        "거래처": ["",         "KT",          ""],
        "출금":  ["3200000",   "88000",        "2800000"],
        "입금":  ["",          "",             ""],
    }).fillna("")


@_skip
def test_summary_sheet_has_합계금액_column():
    """[요약] 시트 헤더에 '합계금액' 컬럼이 있어야 한다."""
    from butler_pc_core.accounting.classifier import classify_df, save_classified

    df = classify_df(_sep_df())
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        out = Path(f.name)
    try:
        save_classified(df, out)
        wb = openpyxl.load_workbook(out)
        ws = wb["요약"]
        headers = [cell.value for cell in ws[1]]
        assert "합계금액" in headers, f"[요약] 시트 헤더에 '합계금액' 없음: {headers}"
        assert "비율" in headers, f"[요약] 시트 헤더에 '비율' 없음: {headers}"
    finally:
        out.unlink(missing_ok=True)


@_skip
def test_summary_sheet_합계금액_values_correct():
    """[요약] 시트 합계금액 합계가 DataFrame _amt 합과 일치해야 한다."""
    from butler_pc_core.accounting.classifier import classify_df, save_classified

    df = classify_df(_sep_df())
    assert "_amt" in df.columns, "_amt 컬럼이 있어야 함"

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        out = Path(f.name)
    try:
        save_classified(df, out)
        wb = openpyxl.load_workbook(out)
        ws = wb["요약"]
        headers = [cell.value for cell in ws[1]]
        amt_col_idx = headers.index("합계금액") + 1  # 1-based

        # 총계 행 (마지막 행)
        last_row = ws.max_row
        total_cell = ws.cell(row=last_row, column=1).value
        assert total_cell == "총계", f"마지막 행이 '총계'가 아님: {total_cell}"

        xlsx_total = ws.cell(row=last_row, column=amt_col_idx).value
        df_total = int(df[df["분류과목"] != "미분류"]["_amt"].sum())
        assert xlsx_total == df_total, (
            f"[요약] 합계금액 총계 불일치: xlsx={xlsx_total}, df={df_total}"
        )
    finally:
        out.unlink(missing_ok=True)


@_skip
def test_summary_sheet_합계금액_number_format_applied():
    """[요약] 시트 합계금액 셀에 콤마 서식이 적용되어야 한다."""
    from butler_pc_core.accounting.classifier import classify_df, save_classified

    df = classify_df(_sep_df())
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        out = Path(f.name)
    try:
        save_classified(df, out)
        wb = openpyxl.load_workbook(out)
        ws = wb["요약"]
        headers = [cell.value for cell in ws[1]]
        amt_col_idx = headers.index("합계금액") + 1

        # 데이터 행(row 2 이상)의 합계금액 셀에 서식 확인
        formatted = False
        for row_idx in range(2, ws.max_row + 1):
            fmt = ws.cell(row=row_idx, column=amt_col_idx).number_format
            if fmt and fmt != "General":
                formatted = True
                break
        assert formatted, "[요약] 합계금액 셀에 숫자 서식이 적용되지 않음"
    finally:
        out.unlink(missing_ok=True)
