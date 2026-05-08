"""test_xlsx_sheet_structure.py — xlsx 시트 순서 + [분류결과] 컬럼 순서 검증."""
from __future__ import annotations

import tempfile
from pathlib import Path

import pytest

try:
    import pandas as pd
    import openpyxl
    _DEPS_OK = True
except ImportError:
    _DEPS_OK = False

_skip = pytest.mark.skipif(not _DEPS_OK, reason="pandas/openpyxl 미설치")


def _classified_df() -> "pd.DataFrame":
    """입출금 분리 컬럼 DataFrame — 분류 성공/실패 혼합."""
    from butler_pc_core.accounting.classifier import classify_df
    return classify_df(pd.DataFrame({
        "번호":  [1, 2, 3],
        "적요":  ["급여 지급", "통신비 납부", "알수없는항목xyz"],
        "거래처": ["", "KT", ""],
        "출금":  ["2500000", "88000", ""],
        "입금":  ["", "", "99999"],
    }).fillna(""))


@_skip
def test_active_sheet_is_요약():
    """xlsx 파일을 열었을 때 활성 시트(첫 번째)가 [요약]이어야 한다."""
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        tmp = Path(f.name)
    try:
        from butler_pc_core.accounting.classifier import save_classified
        save_classified(_classified_df(), tmp)
        wb = openpyxl.load_workbook(tmp)
        assert wb.active is not None, "활성 시트가 None"
        assert wb.active.title == "요약", (
            f"활성 시트가 '요약'이 아님: '{wb.active.title}' — "
            f"시트 순서: {wb.sheetnames}"
        )
        # 시트 순서 검증: [요약, 분류결과, 미분류]
        assert wb.sheetnames == ["요약", "분류결과", "미분류"], (
            f"시트 순서 불일치: {wb.sheetnames}"
        )
    finally:
        tmp.unlink(missing_ok=True)


@_skip
def test_result_sheet_분류과목_in_first_columns():
    """[분류결과] 시트에서 '분류과목' 컬럼이 상위 4열 이내에 위치해야 한다."""
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        tmp = Path(f.name)
    try:
        from butler_pc_core.accounting.classifier import save_classified
        save_classified(_classified_df(), tmp)
        wb = openpyxl.load_workbook(tmp)
        ws = wb["분류결과"]
        headers = [cell.value for cell in ws[1]]
        assert "분류과목" in headers, f"[분류결과] 헤더에 '분류과목' 없음: {headers}"
        idx = headers.index("분류과목")  # 0-based
        assert idx <= 3, (
            f"'분류과목' 컬럼이 너무 오른쪽에 위치 (0-based index={idx}, 기대: ≤3) — "
            f"헤더: {headers}"
        )
        # 신뢰도도 분류과목 바로 다음에 위치해야 함
        assert idx + 1 < len(headers) and headers[idx + 1] == "신뢰도", (
            f"'신뢰도'가 '분류과목' 바로 다음이 아님: {headers}"
        )
    finally:
        tmp.unlink(missing_ok=True)
