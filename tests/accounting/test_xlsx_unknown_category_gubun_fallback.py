"""test_xlsx_unknown_category_gubun_fallback.py — 재검토팀 HOLD 정정 검증.

결함: classifier.py _gubun_label 의 fallback 기본값이 "+"(수익)라서
xlsx 다운로드 [요약]/[분류결과] 시트에서 사전(ACCOUNT_BY_NAME) 미등록
비용 카테고리를 [수익]으로 오표시할 위험이 있었다.

수정: _gubun_label(name, amount) — 미등록 카테고리는 amount 순액 부호로
추론, 사전 등록 카테고리는 acc.sign 우선.

검증 데이터는 통장샘플.csv 원본이 아닌 합성 분류 결과다.
"""
from __future__ import annotations

import tempfile
from pathlib import Path

import pytest

try:
    import openpyxl
    import pandas as pd
    _DEPS_OK = True
except ImportError:
    _DEPS_OK = False

_skip = pytest.mark.skipif(not _DEPS_OK, reason="pandas/openpyxl 미설치")


def _save(rows: list[dict]) -> str:
    """합성 분류 결과 → save_classified → xlsx 경로 반환."""
    from butler_pc_core.accounting.classifier import save_classified

    df = pd.DataFrame(rows)
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        tmp = f.name
    save_classified(df, tmp)
    return tmp


def _summary_gubun(xlsx: str) -> dict[str, str]:
    """[요약] 시트 → {분류과목: 구분} 매핑 (구분 = 컬럼 index 1)."""
    ws = openpyxl.load_workbook(xlsx)["요약"]
    rows = [[c.value for c in r] for r in ws.iter_rows()]
    header = rows[0]
    assert header[0] == "분류과목" and header[1] == "구분", f"[요약] 헤더: {header}"
    return {r[0]: r[1] for r in rows[1:] if r[0] and r[0] != "총계"}


def _classified_gubun(xlsx: str) -> dict[str, str]:
    """[분류결과] 시트 → {분류과목: 구분} 매핑."""
    ws = openpyxl.load_workbook(xlsx)["분류결과"]
    rows = [[c.value for c in r] for r in ws.iter_rows()]
    header = rows[0]
    ci_name = header.index("분류과목")
    ci_gubun = header.index("구분")
    return {r[ci_name]: r[ci_gubun] for r in rows[1:] if r[ci_name]}


# ── 시나리오 1: [요약] 사전 미등록 + 그룹 합계 음수 → "비용" ─────────────────
@_skip
def test_xlsx_summary_unknown_expense():
    xlsx = _save([
        {"적요": "합성거래", "분류과목": "미등록비용X", "신뢰도": 0.85, "_amt": -120000},
    ])
    try:
        assert _summary_gubun(xlsx)["미등록비용X"] == "비용"
    finally:
        Path(xlsx).unlink(missing_ok=True)


# ── 시나리오 2: [분류결과] 사전 미등록 + row._amt 음수 → "비용" ───────────────
@_skip
def test_xlsx_classified_unknown_expense():
    xlsx = _save([
        {"적요": "합성거래", "분류과목": "미등록비용X", "신뢰도": 0.85, "_amt": -120000},
    ])
    try:
        assert _classified_gubun(xlsx)["미등록비용X"] == "비용"
    finally:
        Path(xlsx).unlink(missing_ok=True)


# ── 시나리오 3: 사전 미등록 + amount 양수 → "수익" ([요약]/[분류결과]) ────────
@_skip
def test_xlsx_unknown_revenue():
    xlsx = _save([
        {"적요": "합성거래", "분류과목": "미등록수익X", "신뢰도": 0.85, "_amt": 600000},
    ])
    try:
        assert _summary_gubun(xlsx)["미등록수익X"] == "수익"
        assert _classified_gubun(xlsx)["미등록수익X"] == "수익"
    finally:
        Path(xlsx).unlink(missing_ok=True)


# ── 시나리오 4: 사전 등록 카테고리는 acc.sign 우선 (amount 충돌해도) ─────────
@_skip
def test_xlsx_registered_preserves_acc_sign():
    # 통신비 _amt 를 의도적으로 양수로 충돌시켜도 acc.sign("-") 우선이어야 함
    xlsx = _save([
        {"적요": "합성거래", "분류과목": "매출",   "신뢰도": 0.90, "_amt": 1000000},
        {"적요": "합성거래", "분류과목": "통신비", "신뢰도": 0.90, "_amt": 50000},
    ])
    try:
        summ = _summary_gubun(xlsx)
        clsf = _classified_gubun(xlsx)
        assert summ["매출"] == "수익" and clsf["매출"] == "수익"
        assert summ["통신비"] == "비용", "acc.sign('-') 우선 실패 — 양수 amount에 끌림"
        assert clsf["통신비"] == "비용", "acc.sign('-') 우선 실패 — 양수 amount에 끌림"
    finally:
        Path(xlsx).unlink(missing_ok=True)
