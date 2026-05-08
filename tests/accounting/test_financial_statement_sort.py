"""test_financial_statement_sort.py — 재무제표 표준 정렬 + sign/section 메타데이터 검증."""
from __future__ import annotations

import tempfile
from pathlib import Path

import pytest

try:
    import pandas as pd
    import openpyxl
    _DEPS_OK = True
except ImportError:
    _DEPS_OK = False

_skip = pytest.mark.skipif(not _DEPS_OK, reason="pandas/openpyxl 미설치")


def test_accounts_have_sign_and_section_metadata():
    """모든 ACCOUNTS 항목에 sign ∈ {'+','-'} 와 section ∈ SECTION_ORDER ∪ {'other'} 가 있어야 한다."""
    from butler_pc_core.accounting.account_dict import ACCOUNTS, SECTION_ORDER
    valid_sections = set(SECTION_ORDER.keys()) | {"other"}
    for acc in ACCOUNTS:
        assert acc.sign in ("+", "-"), (
            f"{acc.name}: sign='{acc.sign}' — '+' 또는 '-'이어야 함"
        )
        assert acc.section in valid_sections, (
            f"{acc.name}: section='{acc.section}' — 유효하지 않은 섹션"
        )


def test_지급수수료_has_negative_sign():
    """지급수수료는 비용 계정이므로 sign='-' 이어야 한다."""
    from butler_pc_core.accounting.account_dict import ACCOUNT_BY_NAME
    acc = ACCOUNT_BY_NAME.get("지급수수료")
    assert acc is not None, "지급수수료 계정이 ACCOUNT_BY_NAME에 없음"
    assert acc.sign == "-", f"지급수수료 sign='{acc.sign}' — '-'이어야 함 (사용자 지적 핵심)"


@_skip
def test_expense_account_sign_forced_negative_in_classify_df():
    """입출금 분리 컬럼이 있는 DataFrame에서 비용 계정의 _amt는 음수여야 한다."""
    from butler_pc_core.accounting.classifier import classify_df

    df = classify_df(pd.DataFrame({
        "적요":  ["통신비 납부", "급여 지급", "지급수수료"],
        "거래처": ["KT", "", "세무법인"],
        "출금":  ["88000", "2500000", "500000"],
        "입금":  ["", "", ""],
    }).fillna(""))

    assert "_amt" in df.columns, "_amt 컬럼이 생성되지 않음"

    expense_rows = df[df["분류과목"].isin(["통신비", "직원급여", "지급수수료"])]
    assert not expense_rows.empty, "비용 계정 행이 없음"
    for _, row in expense_rows.iterrows():
        assert row["_amt"] <= 0, (
            f"{row['분류과목']} _amt={row['_amt']} — 비용 계정은 0 이하여야 함"
        )


@_skip
def test_summary_sheet_revenue_appears_first():
    """[요약] 시트 첫 번째 데이터 행의 분류과목이 수익 계정(매출 또는 I_revenue 섹션)이어야 한다."""
    from butler_pc_core.accounting.classifier import classify_df, save_classified
    from butler_pc_core.accounting.account_dict import ACCOUNT_BY_NAME

    df = classify_df(pd.DataFrame({
        "적요":  ["매출 입금", "통신비 납부", "지급수수료"],
        "거래처": ["고객사", "KT", "세무법인"],
        "출금":  ["", "88000", "500000"],
        "입금":  ["5000000", "", ""],
    }).fillna(""))

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        tmp = Path(f.name)
    try:
        save_classified(df, tmp)
        wb = openpyxl.load_workbook(tmp)
        ws = wb["요약"]
        first_data_row = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
        first_category = first_data_row[0]
        acc = ACCOUNT_BY_NAME.get(str(first_category))
        assert acc is not None and acc.section == "I_revenue", (
            f"[요약] 첫 번째 행 계정과목='{first_category}' — I_revenue 섹션이어야 함"
        )
    finally:
        tmp.unlink(missing_ok=True)


@_skip
def test_result_sheet_ordered_by_section():
    """[분류결과] 시트에서 수익 계정(I_revenue) 행이 판관비(IV_sga) 행보다 앞에 위치해야 한다."""
    from butler_pc_core.accounting.classifier import classify_df, save_classified
    from butler_pc_core.accounting.account_dict import ACCOUNT_BY_NAME

    df = classify_df(pd.DataFrame({
        "적요":  ["통신비 납부", "매출 입금", "지급수수료"],
        "거래처": ["KT", "고객사", "세무법인"],
        "출금":  ["88000", "", "500000"],
        "입금":  ["", "5000000", ""],
    }).fillna(""))

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        tmp = Path(f.name)
    try:
        save_classified(df, tmp)
        wb = openpyxl.load_workbook(tmp)
        ws = wb["분류결과"]
        headers = [cell.value for cell in ws[1]]
        assert "분류과목" in headers, f"[분류결과] 헤더에 '분류과목' 없음: {headers}"
        cat_idx = headers.index("분류과목")

        categories = [
            row[cat_idx]
            for row in ws.iter_rows(min_row=2, values_only=True)
            if row[cat_idx] is not None
        ]
        assert categories, "[분류결과] 데이터 행 없음"

        revenue_idx = next(
            (i for i, c in enumerate(categories)
             if ACCOUNT_BY_NAME.get(str(c)) and ACCOUNT_BY_NAME[str(c)].section == "I_revenue"),
            None
        )
        sga_idx = next(
            (i for i, c in enumerate(categories)
             if ACCOUNT_BY_NAME.get(str(c)) and ACCOUNT_BY_NAME[str(c)].section == "IV_sga"),
            None
        )
        assert revenue_idx is not None, "I_revenue 계정 행 없음"
        assert sga_idx is not None, "IV_sga 계정 행 없음"
        assert revenue_idx < sga_idx, (
            f"수익 행(idx={revenue_idx})이 판관비 행(idx={sga_idx})보다 뒤에 위치"
        )
    finally:
        tmp.unlink(missing_ok=True)
