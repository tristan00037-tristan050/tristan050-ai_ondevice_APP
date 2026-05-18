"""test_3sheets.py — 결과 xlsx 3시트 구조 검증."""
from __future__ import annotations

import tempfile
from pathlib import Path

import pytest

try:
    import pandas as pd
    import openpyxl
    _DEPS_OK = True
except ImportError:
    _DEPS_OK = False

_skip = pytest.mark.skipif(not _DEPS_OK, reason="pandas/openpyxl 미설치")


def _mixed_df() -> "pd.DataFrame":
    """분류 성공 + 실패 항목이 섞인 DataFrame."""
    return pd.DataFrame({
        "적요": ["급여 지급", "통신비 납부", "알수없는항목xyz"],
        "거래처": ["", "", ""],
        "금액": [1000000, 50000, 30000],
    })


@_skip
def test_xlsx_has_three_sheets():
    """결과 xlsx에 [분류결과], [요약], [미분류] 3시트가 존재해야 한다."""
    from butler_pc_core.accounting.classifier import classify_df, save_classified

    df = classify_df(_mixed_df())
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        tmp = f.name
    try:
        save_classified(df, tmp)
        wb = openpyxl.load_workbook(tmp)
        assert "분류결과" in wb.sheetnames, f"[분류결과] 시트 없음: {wb.sheetnames}"
        assert "요약" in wb.sheetnames, f"[요약] 시트 없음: {wb.sheetnames}"
        assert "미분류" in wb.sheetnames, f"[미분류] 시트 없음: {wb.sheetnames}"
    finally:
        Path(tmp).unlink(missing_ok=True)


@_skip
def test_summary_sheet_columns():
    """[요약] 시트 헤더가 분류과목, 구분, 건수, 합계금액, 비율이어야 한다."""
    from butler_pc_core.accounting.classifier import classify_df, save_classified

    df = classify_df(_mixed_df())
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        tmp = f.name
    try:
        save_classified(df, tmp)
        wb = openpyxl.load_workbook(tmp)
        ws = wb["요약"]
        header = [cell.value for cell in ws[1]]
        assert header == ["분류과목", "구분", "건수", "합계금액", "비율"], (
            f"[요약] 헤더 불일치: {header}"
        )
    finally:
        Path(tmp).unlink(missing_ok=True)


@_skip
def test_unclassified_sheet_isolated():
    """미분류 항목이 [분류결과]에 없고 [미분류]에만 존재해야 한다."""
    from butler_pc_core.accounting.classifier import classify_df, save_classified

    df = classify_df(_mixed_df())
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        tmp = f.name
    try:
        save_classified(df, tmp)
        wb = openpyxl.load_workbook(tmp)

        # [분류결과] 시트에 미분류 없음
        ws1 = wb["분류결과"]
        headers = [cell.value for cell in ws1[1]]
        acc_col_idx = headers.index("분류과목") + 1  # 1-indexed
        accounts_in_result = [
            ws1.cell(row=r, column=acc_col_idx).value
            for r in range(2, ws1.max_row + 1)
        ]
        assert "미분류" not in accounts_in_result, (
            f"[분류결과] 시트에 미분류 항목 발견: {accounts_in_result}"
        )

        # [미분류] 시트에 행이 존재 (헤더 포함 2행 이상)
        ws3 = wb["미분류"]
        assert ws3.max_row >= 2, "[미분류] 시트에 데이터 행이 없음"
    finally:
        Path(tmp).unlink(missing_ok=True)
