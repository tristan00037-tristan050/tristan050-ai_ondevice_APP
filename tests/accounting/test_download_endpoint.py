"""test_download_endpoint.py — classify → result_id → xlsx 다운로드 엔드포인트 검증."""
from __future__ import annotations

import io
from pathlib import Path

import pytest

try:
    from fastapi.testclient import TestClient
    import openpyxl
    _DEPS_OK = True
except ImportError:
    _DEPS_OK = False

_skip = pytest.mark.skipif(not _DEPS_OK, reason="fastapi/openpyxl 미설치")

_FIXTURE = Path(__file__).parent.parent / "fixtures" / "banks" / "nh_sample.xlsx"


def _get_client():
    import sys, os
    sys.path.insert(0, str(Path(__file__).parent.parent.parent))
    # butler_sidecar를 TestClient로 임포트
    from butler_sidecar import app  # type: ignore
    return TestClient(app, raise_server_exceptions=True)


@_skip
def test_classify_then_download_returns_xlsx():
    """classify → result_id 획득 → xlsx GET → 200 + xlsx Content-Type."""
    client = _get_client()

    # Step 1: classify
    with open(_FIXTURE, "rb") as f:
        resp = client.post(
            "/accounting/classify",
            files={"file": ("nh_sample.xlsx", f,
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
    assert resp.status_code == 200, f"classify failed: {resp.text[:200]}"

    # SSE 스트림에서 result_id 추출
    result_id = None
    import json
    for line in resp.text.splitlines():
        if line.startswith("data:") and '"result_id"' in line:
            result_id = json.loads(line[5:].strip()).get("result_id")
            break
    assert result_id, f"result_id not found in SSE response:\n{resp.text[:500]}"

    # Step 2: download
    dl = client.get(f"/accounting/result/{result_id}/xlsx")
    assert dl.status_code == 200, f"download failed: {dl.text[:200]}"
    assert "spreadsheetml" in dl.headers.get("content-type", ""), (
        f"wrong content-type: {dl.headers.get('content-type')}"
    )
    assert "attachment" in dl.headers.get("content-disposition", "").lower() or \
           "butler_accounting_result.xlsx" in dl.headers.get("content-disposition", ""), (
        f"unexpected content-disposition: {dl.headers.get('content-disposition')}"
    )

    # Step 3: xlsx 파일 유효성 + 3시트 구조
    wb = openpyxl.load_workbook(io.BytesIO(dl.content))
    assert set(wb.sheetnames) >= {"분류결과", "요약", "미분류"}, (
        f"3시트 구조 불완전 — 시트 목록: {wb.sheetnames}"
    )
    ws_result = wb["분류결과"]
    assert ws_result.max_row >= 2, "분류결과 시트 데이터 없음"


@_skip
def test_download_missing_result_id_returns_404():
    """존재하지 않는 result_id로 GET 요청 → 404 응답."""
    client = _get_client()
    resp = client.get("/accounting/result/non-existent-uuid-0000/xlsx")
    assert resp.status_code == 404, f"기대 404, 실제 {resp.status_code}"
